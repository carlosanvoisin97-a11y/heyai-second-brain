import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# SHEET 1: REQUIREMENT LOG
# ============================================================
ws = wb.active
ws.title = "Requirement Log"

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1B3A5C')
p1_fill = PatternFill('solid', fgColor='E74C3C')  # Red - Core
p2_fill = PatternFill('solid', fgColor='F39C12')  # Orange - Important
p3_fill = PatternFill('solid', fgColor='27AE60')  # Green - Nice to have
p1_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
p2_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
p3_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
data_font = Font(name='Arial', size=10)
macro_font = Font(name='Arial', bold=True, size=10, color='1B3A5C')
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

# Headers
headers = [
    ("ID", 10),
    ("Macroarea", 22),
    ("Sottoarea", 24),
    ("Requisito", 65),
    ("Tipo", 14),
    ("Priorità", 10),
    ("Label Priorità", 16),
    ("Note Prototipo / Demo", 45),
    ("Rif. Documento", 16),
    ("Dipendenze", 20),
]

for col_idx, (header, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze top row
ws.freeze_panes = 'A2'
ws.auto_filter.ref = "A1:J1"

# ============================================================
# REQUIREMENT DATA - Extracted from all 24 sections
# ============================================================

reqs = [
    # ── AUTH: Autenticazione & Gestione Accessi (Sez. 5) ──
    ("REQ-001", "AUTH", "Registrazione Azienda", "L'azienda deve potersi registrare inserendo: ragione sociale, partita IVA, email, password", "Funzionale", 1, "Core prototipo", "Form di registrazione con campi obbligatori, validazione P.IVA formato italiano (11 cifre)", "Sez. 5", "—"),
    ("REQ-002", "AUTH", "Autenticazione MFA", "Il sistema deve supportare autenticazione multi-fattore (MFA) al login", "Sicurezza", 2, "Importante", "Nel prototipo: login base email+password. MFA implementato in fase successiva", "Sez. 5", "REQ-001"),
    ("REQ-003", "AUTH", "Accettazione documenti legali", "Durante la registrazione l'utente deve accettare: informativa privacy, condizioni di utilizzo, nomina a responsabile trattamento dati", "Compliance", 1, "Core prototipo", "Checkbox obbligatorie con link a documenti (testi mock per demo)", "Sez. 5", "REQ-001"),
    ("REQ-004", "AUTH", "Login azienda", "L'azienda accede tramite email e password con sessione autenticata", "Funzionale", 1, "Core prototipo", "Login funzionante con sessione base (JWT o session)", "Sez. 5", "REQ-001"),
    ("REQ-005", "AUTH", "Login consulente", "Il consulente del lavoro deve poter accedere con credenziali proprie a una dashboard dedicata", "Funzionale", 1, "Core prototipo", "Ruolo 'consulente' separato con accesso alla propria dashboard", "Sez. 6.1, 17.2", "REQ-001"),
    ("REQ-006", "AUTH", "Gestione ruoli", "Il sistema deve distinguere almeno 3 ruoli: Azienda, Consulente del lavoro, Referente sicurezza", "Funzionale", 1, "Core prototipo", "RBAC base con almeno Azienda e Consulente. Referente sicurezza come P2", "Sez. 5-6", "—"),

    # ── ONB: Onboarding & Profilazione Aziendale (Sez. 9) ──
    ("REQ-007", "ONB", "Tipologia attività", "Menu a tendina per selezione tipologia attività, costruito su attività riconducibili a CCNL Confcommercio", "Funzionale", 1, "Core prototipo", "Dropdown con elenco precaricato di tipologie Confcommercio (10-15 voci principali)", "Sez. 9.1", "REQ-001"),
    ("REQ-008", "ONB", "Orario attività", "L'azienda inserisce: giorni di apertura, ora apertura, ora chiusura", "Funzionale", 1, "Core prototipo", "Multi-select giorni + time picker apertura/chiusura", "Sez. 9.2", "REQ-007"),
    ("REQ-009", "ONB", "Fascia fatturato annuo", "Selezione per fasce: fino a 250K€, 250K-500K€, 500K-1M€, oltre 1M€", "Funzionale", 1, "Core prototipo", "Radio button o dropdown con le 4 fasce indicate", "Sez. 9.3", "REQ-007"),
    ("REQ-010", "ONB", "Dati aggiuntivi settoriali", "Modulo opzionale per dati specifici di settore: ristorazione (coperti, tavoli), ricettivo (camere, posti letto), negozi (superficie), servizi persona (postazioni)", "Funzionale", 2, "Importante", "Nel prototipo: form condizionale che mostra campi diversi in base alla tipologia selezionata in REQ-007", "Sez. 10.2", "REQ-007"),
    ("REQ-011", "ONB", "Wizard onboarding guidato", "Flusso guidato step-by-step per completare la configurazione iniziale al primo accesso", "UX", 1, "Core prototipo", "Wizard 3-4 step: dati azienda → tipologia → orario → fatturato → professionisti", "Sez. 9", "REQ-001"),

    # ── PROF: Gestione Professionisti di Riferimento (Sez. 6) ──
    ("REQ-012", "PROF", "Consulente del lavoro obbligatorio", "L'azienda deve indicare obbligatoriamente: nome consulente, studio, email del consulente del lavoro", "Funzionale", 1, "Core prototipo", "Campi obbligatori nell'onboarding. L'email genera invito al consulente", "Sez. 6.1", "REQ-011"),
    ("REQ-013", "PROF", "Dashboard consulente multi-azienda", "Il consulente accede a una dashboard dedicata con vista su tutte le aziende che lo hanno collegato", "Funzionale", 1, "Core prototipo", "Dashboard consulente con lista aziende collegate e indicatori aggregati", "Sez. 6.1", "REQ-005, REQ-012"),
    ("REQ-014", "PROF", "Referente sicurezza opzionale", "L'azienda può indicare: RSPP esterno, consulente sicurezza, società specializzata (nome, email)", "Funzionale", 2, "Importante", "Campo opzionale nell'onboarding. Se compilato, abilita notifiche sicurezza dedicate", "Sez. 6.2", "REQ-011"),
    ("REQ-015", "PROF", "Separazione competenze CdL vs sicurezza", "Il sistema deve distinguere che il consulente del lavoro non necessariamente segue l'area sicurezza, instradando le notifiche al referente corretto", "Funzionale", 2, "Importante", "Logica di routing notifiche: area lavoro → CdL, area sicurezza → referente sicurezza (se presente) o azienda", "Sez. 6.2, 16", "REQ-012, REQ-014"),

    # ── DATA: Alimentazione & Estrazione Dati (Sez. 7-8) ──
    ("REQ-016", "DATA", "Upload cedolini PDF", "L'azienda deve poter caricare cedolini paga in formato PDF tramite upload diretto nella piattaforma", "Funzionale", 1, "Core prototipo", "Upload drag&drop di PDF singoli o multipli. Per il prototipo: file accettato e simulazione estrazione con dati mock", "Sez. 7", "REQ-004"),
    ("REQ-017", "DATA", "Upload cedolini via email", "Il sistema deve supportare l'inoltro cedolini a un indirizzo email dedicato (es. azienda@upload.presidia.it) con importazione automatica", "Funzionale", 3, "Nice-to-have", "Non in scope prototipo. Richiede infrastruttura email parsing", "Sez. 7", "REQ-016"),
    ("REQ-018", "DATA", "Periodicità upload", "Il caricamento è mensile e obbligatorio come unica attività periodica richiesta all'azienda", "Business Rule", 1, "Core prototipo", "Dashboard mostra ultimo caricamento + reminder se mancante da >30gg", "Sez. 7", "REQ-016"),
    ("REQ-019", "DATA", "Estrazione dati contrattuali", "Il sistema estrae dai cedolini: livello, tipo contratto, data assunzione, eventuale scadenza, part-time/full-time", "Funzionale", 1, "Core prototipo", "Per prototipo: dati precompilati/mockati che simulano l'estrazione. Mapping su struttura dati definita", "Sez. 8.1", "REQ-016"),
    ("REQ-020", "DATA", "Estrazione dati retributivi", "Il sistema estrae: retribuzione lorda, straordinari, trasferte, rimborsi spese, voci variabili", "Funzionale", 1, "Core prototipo", "Dati mock strutturati coerenti tra loro per alimentare dashboard", "Sez. 8.2", "REQ-016"),
    ("REQ-021", "DATA", "Estrazione dati organizzativi", "Il sistema estrae: ferie maturate, ferie godute, ferie residue, ore lavorate", "Funzionale", 1, "Core prototipo", "Dati mock con valori realistici per visualizzazione dashboard", "Sez. 8.3", "REQ-016"),
    ("REQ-022", "DATA", "Motore OCR/AI estrazione", "Il sistema deve includere un motore di estrazione automatica dei dati dai PDF cedolini (OCR + NLP)", "Funzionale", 2, "Importante", "Fuori scope prototipo. Si usano dati mockati. Da sviluppare con pipeline AI dedicata", "Sez. 7-8", "REQ-016"),
    ("REQ-023", "DATA", "Aggiornamento dashboard post-upload", "Dopo il caricamento dei cedolini, il sistema importa i documenti, estrae i dati e aggiorna la dashboard", "Funzionale", 1, "Core prototipo", "Nel prototipo: caricamento simula elaborazione e popola dashboard con dati mock", "Sez. 7", "REQ-016, REQ-019"),

    # ── PRIV: Architettura Privacy (Sez. 4) ──
    ("REQ-024", "PRIV", "Archivio A – Aziendale", "Archivio contenente: nominativi dipendenti, documenti caricati, collegamento dipendente-ID pseudonimo. Accesso riservato solo all'azienda", "Privacy", 2, "Importante", "Per prototipo: struttura dati separata, accesso filtrato per ruolo. Separazione logica anche se non fisica", "Sez. 4.1", "REQ-006"),
    ("REQ-025", "PRIV", "Archivio B – Pseudonimizzato", "Archivio contenente: ID anonimo dipendente, dati contrattuali estratti, dati retributivi minimi, indicatori organizzativi. Il motore di analisi opera SOLO su questo archivio", "Privacy", 2, "Importante", "Separazione logica nel data model: tabella con soli ID pseudonimi e dati numerici", "Sez. 4.2", "REQ-024"),
    ("REQ-026", "PRIV", "Ricomposizione finale", "L'analisi avviene su dati pseudonimizzati; il sistema ricollega poi l'esito al dipendente nell'archivio aziendale solo per la visualizzazione lato azienda", "Privacy", 2, "Importante", "Join logico A↔B eseguito solo al momento della visualizzazione dashboard azienda", "Sez. 4.3", "REQ-024, REQ-025"),
    ("REQ-027", "PRIV", "Privacy by design", "L'intera architettura deve essere progettata in logica privacy by design secondo GDPR", "Non-Funzionale", 2, "Importante", "Nel prototipo: documentare l'architettura privacy. Implementazione completa in produzione", "Sez. 4", "—"),

    # ── DASH: Dashboard Aziendale (Sez. 11) ──
    ("REQ-028", "DASH", "Struttura del personale", "Visualizzazione: numero dipendenti, full-time/part-time, apprendisti, tempo determinato", "Funzionale", 1, "Core prototipo", "Card/widget con conteggi e distribuzione. Dati mock per 10-15 dipendenti fittizi", "Sez. 11.1", "REQ-019"),
    ("REQ-029", "DASH", "Costo del personale", "Visualizzazione: costo mensile, proiezione annua, costo per singolo dipendente", "Funzionale", 1, "Core prototipo", "Widget con valore mensile, proiezione ×12, e tabella costo/dipendente", "Sez. 11.2", "REQ-020"),
    ("REQ-030", "DASH", "Ore lavorate e straordinari", "Visualizzazione: ore complessive, ore straordinarie, dipendenti con straordinari", "Funzionale", 1, "Core prototipo", "Grafico a barre mensile + highlight dipendenti con straordinari", "Sez. 11.3", "REQ-020, REQ-021"),
    ("REQ-031", "DASH", "Ferie", "Visualizzazione: ferie maturate, ferie godute, ferie residue (aggregate e per dipendente)", "Funzionale", 1, "Core prototipo", "Widget con totali + tabella dettaglio. Semaforo su ferie residue eccessive", "Sez. 11.4", "REQ-021"),
    ("REQ-032", "DASH", "Voci retributive particolari", "Visualizzazione: trasferte, rimborsi spese, competenze variabili e frequenza della loro presenza nel tempo", "Funzionale", 1, "Core prototipo", "Tabella con trend mensile delle voci particolari", "Sez. 11.5", "REQ-020"),
    ("REQ-033", "DASH", "Scadenze contrattuali", "Visualizzazione: contratti a termine in scadenza, apprendistati in scadenza con timeline", "Funzionale", 1, "Core prototipo", "Lista ordinata per prossimità scadenza con semaforo", "Sez. 11.6", "REQ-019"),
    ("REQ-034", "DASH", "Dinamica dell'organico", "Visualizzazione: nuovi ingressi, cessazioni, variazione organico nel periodo", "Funzionale", 1, "Core prototipo", "Widget con delta organico mese su mese", "Sez. 11.7", "REQ-019"),
    ("REQ-035", "DASH", "Fotografia dell'attività", "Se configurata, mostra: tipologia attività, ore apertura, dimensione attività, sintesi organizzativa", "Funzionale", 1, "Core prototipo", "Sezione riepilogativa che riprende dati onboarding + sintesi da osservatorio", "Sez. 11.8", "REQ-007, REQ-008, REQ-009"),
    ("REQ-036", "DASH", "Risposta immediata post-caricamento", "L'azienda deve vedere risultati utili IMMEDIATAMENTE dopo il caricamento dei cedolini", "UX", 1, "Core prototipo", "Feedback visivo: progress bar → dashboard aggiornata. Tempo percepito < 5 sec", "Sez. 11", "REQ-023"),

    # ── SEM: Sistema Semaforo (Sez. 12) ──
    ("REQ-037", "SEM", "Indicatore Verde", "🟢 Situazione stabile — nessuna azione richiesta", "Funzionale", 1, "Core prototipo", "Badge/icona verde su ogni indicatore monitorato", "Sez. 12", "REQ-028-035"),
    ("REQ-038", "SEM", "Indicatore Giallo", "🟡 Elemento da monitorare — attenzione preventiva raccomandata", "Funzionale", 1, "Core prototipo", "Badge giallo + tooltip con spiegazione", "Sez. 12", "REQ-028-035"),
    ("REQ-039", "SEM", "Indicatore Rosso", "🔴 Situazione che merita attenzione — azione suggerita (confronto con consulente)", "Funzionale", 1, "Core prototipo", "Badge rosso + messaggio di invito a confronto con consulente", "Sez. 12", "REQ-028-035"),
    ("REQ-040", "SEM", "Indicatore Blu", "🔵 Decisione organizzativa registrata dall'azienda — stato definito", "Funzionale", 1, "Core prototipo", "Badge blu quando l'azienda ha registrato una scelta consapevole", "Sez. 12", "REQ-028-035"),
    ("REQ-041", "SEM", "Soglie semaforo configurabili", "Le soglie che determinano il passaggio da verde a giallo a rosso devono essere definite per ogni indicatore", "Business Rule", 2, "Importante", "Per prototipo: soglie hardcoded per indicatori principali (ferie, straordinari, scadenze). Da rendere configurabili", "Sez. 12", "REQ-037-040"),

    # ── VOCI: Gestione Voci Retributive Particolari (Sez. 13) ──
    ("REQ-042", "VOCI", "Rilevazione voci particolari", "Il sistema rileva e visualizza la presenza di: trasferte, rimborsi, voci non imponibili, altre competenze variabili", "Funzionale", 1, "Core prototipo", "Sezione dedicata con elenco voci rilevate dai dati mock", "Sez. 13", "REQ-020"),
    ("REQ-043", "VOCI", "Messaggio prudenziale", "Visualizzazione messaggio tipo: 'Sono state rilevate alcune voci retributive che non incidono integralmente sull'imponibile contributivo o fiscale. Può essere utile verificare la gestione di tali voci con il consulente del lavoro.'", "Funzionale", 1, "Core prototipo", "Messaggio fisso con tono informativo, mai consulenziale", "Sez. 13", "REQ-042"),
    ("REQ-044", "VOCI", "Opzione: Verifica con CdL", "L'azienda può selezionare 'Verifica con il consulente del lavoro' per ogni voce rilevata", "Funzionale", 1, "Core prototipo", "Bottone/checkbox che genera notifica al consulente", "Sez. 13", "REQ-042, REQ-012"),
    ("REQ-045", "VOCI", "Opzione: Voci documentate", "L'azienda può selezionare 'Voci supportate da documentazione' — il sistema registra la scelta organizzativa (semaforo blu)", "Funzionale", 1, "Core prototipo", "Checkbox + registrazione nel registro decisionale + cambio semaforo a blu", "Sez. 13", "REQ-042, REQ-040"),

    # ── MON: Aree di Monitoraggio (Sez. 14) ──
    ("REQ-046", "MON", "Inquadramento e mansioni", "Confronto tra livello in cedolino, mansione effettiva e declaratorie CCNL Confcommercio", "Funzionale", 2, "Importante", "Per prototipo: campo mansione editabile + confronto con tabella declaratorie CCNL precaricata", "Sez. 14.1", "REQ-019"),
    ("REQ-047", "MON", "Contratti a termine e apprendistato", "Monitoraggio di: scadenze, proroghe, durata complessiva con alert su limiti normativi", "Funzionale", 1, "Core prototipo", "Timeline scadenze + alert se prossimità scadenza o superamento limiti durata", "Sez. 14.2", "REQ-019, REQ-033"),
    ("REQ-048", "MON", "Ferie – 2 settimane anno maturazione", "Verifica che almeno 2 settimane di ferie siano godute nell'anno di maturazione", "Business Rule", 1, "Core prototipo", "Calcolo automatico su dati mock: ferie godute vs 2 settimane. Semaforo se non rispettato", "Sez. 14.3", "REQ-021, REQ-031"),
    ("REQ-049", "MON", "Ferie – residuo entro 18 mesi", "Verifica godimento ferie residue entro 18 mesi dalla maturazione", "Business Rule", 1, "Core prototipo", "Alert se ferie residue > 18 mesi dalla maturazione", "Sez. 14.3", "REQ-021, REQ-031"),
    ("REQ-050", "MON", "Straordinari – monitoraggio", "Monitoraggio di: frequenza straordinari, andamento mensile, proiezione annua", "Funzionale", 1, "Core prototipo", "Grafico trend mensile + proiezione annua su base dati correnti", "Sez. 14.4", "REQ-020, REQ-030"),
    ("REQ-051", "MON", "Part-time pattern", "Rilevazione pattern: part-time + assenza ore supplementari, part-time + voci variabili ricorrenti, part-time + compensi non imponibili", "Funzionale", 2, "Importante", "Logica di pattern matching su dati strutturati. Per prototipo: almeno 1 pattern visibile su dati mock", "Sez. 14.5", "REQ-019, REQ-020"),
    ("REQ-052", "MON", "Turnover", "Monitoraggio: cessazioni, nuovi ingressi, stabilità dell'organico", "Funzionale", 1, "Core prototipo", "KPI turnover rate + visualizzazione trend", "Sez. 14.6", "REQ-034"),
    ("REQ-053", "MON", "Analisi incrociate avanzate (consulente)", "Le analisi incrociate avanzate tra più indicatori possono essere riservate alla dashboard consulente", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Placeholder nella UI consulente", "Sez. 14.6", "REQ-013"),

    # ── OSS: Osservatorio Organizzativo Attività (Sez. 10) ──
    ("REQ-054", "OSS", "Incrocio dati attività-organico", "Il sistema incrocia: tipologia attività, orario apertura, dimensione attività, numero dipendenti, monte ore complessivo, PT/FT, straordinari, costo personale", "Funzionale", 2, "Importante", "Per prototipo: logica semplificata con output descrittivo basato su parametri principali", "Sez. 10.3", "REQ-007-010, REQ-019-021"),
    ("REQ-055", "OSS", "Output lettura organizzativa", "Output testuale NON di conformità. Esempi: 'La struttura dell'organico appare contenuta rispetto alla dimensione dell'attività'", "Funzionale", 2, "Importante", "Messaggi pre-definiti selezionati in base a soglie su parametri incrociati", "Sez. 10.4", "REQ-054"),
    ("REQ-056", "OSS", "Avvertenza metodologica", "Il modulo deve SEMPRE visualizzare la nota: 'Gli indicatori sono elaborati sulla base di informazioni organizzative inserite dall'azienda e dei dati contenuti nei cedolini. Non rappresentano una valutazione di conformità normativa.'", "Compliance", 1, "Core prototipo", "Disclaimer fisso non rimovibile in testa o in coda alla sezione osservatorio", "Sez. 10.5", "REQ-054"),
    ("REQ-057", "OSS", "NON conformità normativa", "Il sistema NON deve mai parlare di conformità/non conformità normativa. Solo lettura organizzativa", "Compliance", 1, "Core prototipo", "Vincolo di design: nessun elemento UI o copy deve usare parole come 'conforme', 'violazione', 'illecito'", "Sez. 1, 10", "—"),

    # ── SIC: Modulo Sicurezza sul Lavoro (Sez. 15) ──
    ("REQ-058", "SIC", "Checklist adempimenti sicurezza", "L'azienda visualizza gli adempimenti principali: DVR, nomina RSPP, formazione lavoratori, addetti antincendio e primo soccorso, sorveglianza sanitaria/medico competente, formazione datore di lavoro", "Funzionale", 1, "Core prototipo", "Lista di 6 adempimenti principali con stato selezionabile", "Sez. 15.3", "REQ-004"),
    ("REQ-059", "SIC", "Stato adempimento", "Per ciascun adempimento l'azienda indica: presente, non presente, non applicabile", "Funzionale", 1, "Core prototipo", "Tri-state selector (presente/non presente/N.A.) per ogni voce checklist", "Sez. 15.2", "REQ-058"),
    ("REQ-060", "SIC", "Upload documentazione sicurezza", "L'azienda può eventualmente caricare la documentazione a supporto di ogni adempimento", "Funzionale", 2, "Importante", "Upload file per adempimento. Nel prototipo: UI presente, funzionalità base", "Sez. 15.2", "REQ-058"),
    ("REQ-061", "SIC", "Semaforo sicurezza", "Il sistema genera il semaforo sulla base dei flag aziendali (presente/non presente/N.A.)", "Funzionale", 1, "Core prototipo", "Semaforo automatico: tutto presente = verde, mancanze = giallo/rosso", "Sez. 15.4", "REQ-059, REQ-037-040"),
    ("REQ-062", "SIC", "Notifiche sicurezza al referente", "Se esiste un referente sicurezza → notifica a lui. Se non esiste → notifica all'azienda", "Funzionale", 1, "Core prototipo", "Logica condizionale su presenza referente. Per demo: notifica simulata (log/UI)", "Sez. 15.5", "REQ-014, REQ-015"),
    ("REQ-063", "SIC", "Contatto società specializzata", "Possibilità di richiedere contatto con società specializzata o convenzionata per sicurezza", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Placeholder 'Richiedi consulenza' disabilitato", "Sez. 15.5", "REQ-062"),
    ("REQ-064", "SIC", "Autonomia modulo", "Il modulo sicurezza è autonomo e funziona indipendentemente dal caricamento cedolini", "Architettura", 1, "Core prototipo", "Modulo accessibile e funzionante anche senza dati cedolini caricati", "Sez. 15.1", "—"),

    # ── NOT: Sistema Notifiche (Sez. 16) ──
    ("REQ-065", "NOT", "Notifiche area lavoro → CdL", "Per indicatori su contratti, ferie, straordinari, voci retributive, turnover: la notifica parte PRIMA al consulente del lavoro", "Funzionale", 1, "Core prototipo", "Per demo: notifica simulata in dashboard consulente (inbox/alert). Email reale in produzione", "Sez. 16.1", "REQ-012, REQ-013"),
    ("REQ-066", "NOT", "Notifiche area sicurezza → referente", "Notifica prioritaria al referente sicurezza (se presente), altrimenti all'azienda", "Funzionale", 1, "Core prototipo", "Per demo: routing condizionale della notifica con log visibile", "Sez. 16.2", "REQ-014, REQ-062"),
    ("REQ-067", "NOT", "Fotografia immediata per azienda", "Nelle more della notifica al professionista, l'azienda vede COMUNQUE la fotografia organizzativa immediata", "UX", 1, "Core prototipo", "Dashboard sempre visibile indipendentemente dallo stato notifiche. Nessun blocco 'in attesa'", "Sez. 16.3", "REQ-028-036"),
    ("REQ-068", "NOT", "Differenziazione notifiche per area", "Le notifiche devono essere differenziate e instradate per area (lavoro vs sicurezza)", "Funzionale", 1, "Core prototipo", "Tag/categoria su ogni notifica per routing corretto", "Sez. 16", "REQ-065, REQ-066"),

    # ── CONS: Dashboard Consulente (Sez. 17) ──
    ("REQ-069", "CONS", "Vista multi-azienda", "Il consulente vede un elenco di tutte le aziende collegate con indicatori sintetici", "Funzionale", 1, "Core prototipo", "Lista aziende con semaforo aggregato per ciascuna", "Sez. 6.1, 17.2", "REQ-013"),
    ("REQ-070", "CONS", "Analisi integrate", "Il consulente vede analisi integrate non visibili all'azienda", "Funzionale", 2, "Importante", "Per prototipo: sezione placeholder con 1-2 analisi incrociate mock", "Sez. 17.2", "REQ-069"),
    ("REQ-071", "CONS", "Pattern organizzativi avanzati", "Il consulente vede pattern organizzativi più avanzati (es. correlazioni tra indicatori)", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo", "Sez. 17.2", "REQ-070"),
    ("REQ-072", "CONS", "Alert preventivi", "Il consulente riceve alert preventivi prima che la situazione diventi critica", "Funzionale", 2, "Importante", "Per prototipo: alert pre-configurati su soglie principali (ferie, scadenze)", "Sez. 17.2", "REQ-065"),
    ("REQ-073", "CONS", "Letture incrociate", "Il consulente può effettuare letture incrociate tra diversi indicatori di diverse aziende", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo", "Sez. 17.2", "REQ-069"),
    ("REQ-074", "CONS", "Inbox notifiche consulente", "Il consulente ha una sezione dedicata per visualizzare tutte le notifiche/segnalazioni ricevute", "Funzionale", 1, "Core prototipo", "Lista notifiche con filtro per azienda, area, priorità", "Sez. 16, 17.2", "REQ-065"),

    # ── NORM: Sezione Informazioni Normative (Sez. 18) ──
    ("REQ-075", "NORM", "Q&A CCNL e normativa", "Sezione che risponde a domande su CCNL, normativa del lavoro, istituti contrattuali", "Funzionale", 2, "Importante", "Per prototipo: FAQ pre-caricate su principali istituti CCNL Confcommercio (10-20 domande)", "Sez. 18", "—"),
    ("REQ-076", "NORM", "Formulazione risposte normative", "Le risposte devono SEMPRE essere formulate come: 'Il CCNL prevede che…' / 'La normativa prevede che…'. MAI in forma consulenziale", "Compliance", 1, "Core prototipo", "Template di risposta vincolante. Nessuna frase che inizi con 'dovresti', 'ti consiglio', ecc.", "Sez. 18", "REQ-075, REQ-057"),
    ("REQ-077", "NORM", "Rimando al consulente", "Quando la domanda richiede valutazione applicativa, il sistema risponde: 'Per eventuali valutazioni specifiche è opportuno confrontarsi con il consulente del lavoro'", "Compliance", 1, "Core prototipo", "Messaggio standard sempre presente quando la domanda esce dal perimetro informativo", "Sez. 18", "REQ-075"),
    ("REQ-078", "NORM", "Motore AI Q&A (futuro)", "Evoluzione futura: motore AI conversazionale alimentato da base dati CCNL e normativa", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Per demo: ricerca keyword su FAQ statiche", "Sez. 18, 19", "REQ-075"),

    # ── AVA: Avatar Digitale (Sez. 19) ──
    ("REQ-079", "AVA", "Avatar guida piattaforma", "Avatar digitale con funzione di: guida della piattaforma, spiegazione dei risultati, supporto informativo normativo", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Eventuale chatbot base in fase successiva", "Sez. 19", "—"),
    ("REQ-080", "AVA", "Personalizzazione avatar consulente", "Il consulente può personalizzare l'avatar con: immagine, voce, nome. Servizio opzionale a suo carico", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Feature premium futura", "Sez. 19", "REQ-079"),

    # ── REG: Registro Decisionale e Storico (Sez. 20) ──
    ("REQ-081", "REG", "Registro decisionale", "Il sistema registra nel tempo: indicatori rilevati, decisioni organizzative registrate dall'azienda", "Funzionale", 1, "Core prototipo", "Tabella log con timestamp, tipo indicatore, decisione presa, utente", "Sez. 20", "REQ-045"),
    ("REQ-082", "REG", "Storico andamento personale", "Tracciamento storico: andamento del personale, evoluzione costo del lavoro, dinamiche principali", "Funzionale", 2, "Importante", "Per prototipo: struttura dati che accumula dati mese su mese. Visualizzazione trend base", "Sez. 20", "REQ-028-034"),
    ("REQ-083", "REG", "Memoria organizzativa", "PresidIA costruisce nel tempo una memoria organizzativa dell'azienda consultabile", "Funzionale", 2, "Importante", "Architettura dati che preserva storico. UI di consultazione base", "Sez. 20", "REQ-081, REQ-082"),
    ("REQ-084", "REG", "Portabilità storico a nuovo consulente", "Se cambia il consulente, il nuovo professionista può accedere allo storico organizzativo autorizzato dall'azienda", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Richiede gestione autorizzazioni e trasferimento", "Sez. 20", "REQ-083"),

    # ── REP: Report Annuale (Sez. 21) ──
    ("REQ-085", "REP", "Generazione report annuale", "Il sistema genera un report annuale con: struttura personale, costo personale, turnover, ferie, straordinari, indicatori emersi, decisioni registrate", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Template PDF/Word da implementare quando ci sono dati reali su 12 mesi", "Sez. 21", "REQ-081-083"),

    # ── NFR: Requisiti Non-Funzionali ──
    ("REQ-086", "NFR", "Semplicità operativa", "Il sistema deve essere utilizzabile facilmente anche da PMI senza struttura HR interna", "UX", 1, "Core prototipo", "UI minimale, linguaggio semplice, wizard guidati, nessun gergo tecnico HR", "Sez. 3.5", "—"),
    ("REQ-087", "NFR", "Prudenza giuridica", "PresidIA non certifica conformità normativa e non formula pareri professionali. Questo vincolo deve permeare OGNI output del sistema", "Compliance", 1, "Core prototipo", "Audit di ogni copy/messaggio del sistema per verificare assenza di toni consulenziali", "Sez. 3.6", "REQ-057, REQ-076"),
    ("REQ-088", "NFR", "Responsiveness", "La piattaforma deve essere accessibile da desktop e mobile (responsive design)", "UX", 2, "Importante", "Per prototipo: desktop-first. Mobile responsive come P2", "—", "—"),
    ("REQ-089", "NFR", "Performance caricamento", "Dashboard deve caricarsi in < 3 secondi. Upload cedolini deve dare feedback entro 5 secondi", "Performance", 2, "Importante", "Con dati mock: performance garantita. Da monitorare con dati reali", "—", "REQ-036"),
    ("REQ-090", "NFR", "Scalabilità", "Il sistema deve supportare la gestione di più aziende per consulente e più dipendenti per azienda", "Architettura", 2, "Importante", "Architettura multi-tenant fin dal design iniziale", "Sez. 22", "—"),
    ("REQ-091", "NFR", "Sicurezza dati", "Crittografia dati a riposo e in transito. Backup periodici. Log accessi", "Sicurezza", 2, "Importante", "Per prototipo: HTTPS + autenticazione. Crittografia completa in produzione", "Sez. 4", "REQ-027"),
    ("REQ-092", "NFR", "Centralità del consulente", "Il sistema deve SEMPRE rafforzare il ruolo del consulente, mai sostituirlo o sminuirlo", "Business Rule", 1, "Core prototipo", "Principio di design: ogni alert rimanda al consulente, mai azioni autonome", "Sez. 3.3, 3.4", "—"),
    ("REQ-093", "NFR", "Localizzazione", "Interfaccia interamente in italiano. Formati data, valuta, numeri italiani", "UX", 1, "Core prototipo", "UI in italiano, formato dd/mm/yyyy, € come valuta, separatore migliaia .", "—", "—"),
    ("REQ-094", "NFR", "CCNL scope iniziale", "Il prototipo copre esclusivamente il CCNL Confcommercio. L'architettura deve prevedere estensibilità ad altri CCNL", "Architettura", 1, "Core prototipo", "Struttura dati parametrica per CCNL (tabella declaratorie, soglie, istituti) anche se popolata solo Confcommercio", "Sez. 9.1, 10, 23", "—"),

    # ── POS: Posizionamento e Identità (Sez. 1, 23, 24) ──
    ("REQ-095", "POS", "Identità negativa", "Il sistema NON è: un software paghe, un gestionale HR tradizionale, un sistema di consulenza automatica, uno strumento di controllo del consulente", "Business Rule", 1, "Core prototipo", "Vincolo di design e comunicazione: nessuna feature che simuli paghe, HR gestionale, o consulenza automatica", "Sez. 1", "—"),
    ("REQ-096", "POS", "Identità positiva", "Il sistema È: uno strumento di consapevolezza organizzativa, un sistema di monitoraggio preventivo, un facilitatore del rapporto azienda-consulente", "Business Rule", 1, "Core prototipo", "Ogni feature e copy deve essere allineato a questa identità", "Sez. 1", "—"),

    # ── PITCH: Requisiti aggiuntivi dal pitch deck (Sez. PDF) ──
    ("REQ-097", "PITCH", "Simulatore costi del personale", "Simulatore a fini previsionali per proiezione costi del personale (cosa succede se assumo X, se aumento Y)", "Funzionale", 3, "Nice-to-have", "Fuori scope prototipo. Feature avanzata post-launch", "PDF pag. 5", "REQ-029"),
    ("REQ-098", "PITCH", "Gestione scadenze e adempimenti", "Modulo dedicato alla gestione centralizzata di scadenze e adempimenti normativi", "Funzionale", 2, "Importante", "Per prototipo: scadenze contrattuali già coperte. Adempimenti sicurezza già in scope. Calendario unificato come P2", "PDF pag. 5", "REQ-033, REQ-058"),
    ("REQ-099", "PITCH", "Scalabilità a altri CCNL", "Visione strategica: estensione del modello ad altri CCNL oltre Confcommercio", "Architettura", 3, "Nice-to-have", "Architettura parametrica già prevista in REQ-094. Contenuti CCNL aggiuntivi post-launch", "PDF pag. 10", "REQ-094"),
    ("REQ-100", "PITCH", "Estensione ambiti normativi", "Visione strategica: estensione a sicurezza, welfare, ESG, parità, formazione", "Business Rule", 3, "Nice-to-have", "Moduli futuri. Architettura modulare che lo consenta", "PDF pag. 10", "—"),
    ("REQ-101", "PITCH", "Hub nazionale imprese-professionisti", "Visione strategica: diventare la piattaforma di riferimento per la gestione informativa e preventiva dei rapporti di lavoro in Italia", "Business Rule", 3, "Nice-to-have", "Visione a lungo termine. Non impatta prototipo", "PDF pag. 10", "—"),
]

# Write data
for row_idx, req in enumerate(reqs, 2):
    for col_idx, value in enumerate(req, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = wrap_align
        cell.border = thin_border
    
    # Style priority column
    priority = req[5]  # Priority value
    p_cell = ws.cell(row=row_idx, column=6)
    p_cell.alignment = center_align
    label_cell = ws.cell(row=row_idx, column=7)
    
    if priority == 1:
        p_cell.fill = p1_fill
        p_cell.font = p1_font
        label_cell.fill = PatternFill('solid', fgColor='FADBD8')
    elif priority == 2:
        p_cell.fill = p2_fill
        p_cell.font = p2_font
        label_cell.fill = PatternFill('solid', fgColor='FDEBD0')
    else:
        p_cell.fill = p3_fill
        p_cell.font = p3_font
        label_cell.fill = PatternFill('solid', fgColor='D5F5E3')
    
    # Bold macroarea
    ws.cell(row=row_idx, column=2).font = macro_font

# ============================================================
# SHEET 2: RIEPILOGO MACROAREE
# ============================================================
ws2 = wb.create_sheet("Riepilogo Macroaree")

# Count by macroarea and priority
from collections import defaultdict
counts = defaultdict(lambda: {1: 0, 2: 0, 3: 0, 'total': 0})
for req in reqs:
    area = req[1]
    pri = req[5]
    counts[area][pri] += 1
    counts[area]['total'] += 1

# Headers
summary_headers = ["Macroarea", "Descrizione", "P1 Core", "P2 Importante", "P3 Nice-to-have", "Totale"]
for col_idx, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 10

area_descriptions = {
    "AUTH": "Autenticazione & Gestione Accessi",
    "ONB": "Onboarding & Profilazione Aziendale",
    "PROF": "Gestione Professionisti di Riferimento",
    "DATA": "Alimentazione & Estrazione Dati",
    "PRIV": "Architettura Privacy",
    "DASH": "Dashboard Aziendale",
    "SEM": "Sistema Semaforo",
    "VOCI": "Gestione Voci Retributive Particolari",
    "MON": "Aree di Monitoraggio",
    "OSS": "Osservatorio Organizzativo Attività",
    "SIC": "Modulo Sicurezza sul Lavoro",
    "NOT": "Sistema Notifiche",
    "CONS": "Dashboard Consulente",
    "NORM": "Sezione Informazioni Normative",
    "AVA": "Avatar Digitale",
    "REG": "Registro Decisionale e Storico",
    "REP": "Report Annuale",
    "NFR": "Requisiti Non-Funzionali",
    "POS": "Posizionamento e Identità Prodotto",
    "PITCH": "Requisiti da Pitch Deck",
}

# Maintain order
area_order = ["AUTH", "ONB", "PROF", "DATA", "PRIV", "DASH", "SEM", "VOCI", "MON", "OSS", "SIC", "NOT", "CONS", "NORM", "AVA", "REG", "REP", "NFR", "POS", "PITCH"]

row = 2
for area in area_order:
    if area in counts:
        c = counts[area]
        ws2.cell(row=row, column=1, value=area).font = macro_font
        ws2.cell(row=row, column=2, value=area_descriptions.get(area, area)).font = data_font
        
        p1_c = ws2.cell(row=row, column=3, value=c[1])
        p1_c.font = data_font
        if c[1] > 0:
            p1_c.fill = PatternFill('solid', fgColor='FADBD8')
        
        p2_c = ws2.cell(row=row, column=4, value=c[2])
        p2_c.font = data_font
        if c[2] > 0:
            p2_c.fill = PatternFill('solid', fgColor='FDEBD0')
        
        p3_c = ws2.cell(row=row, column=5, value=c[3])
        p3_c.font = data_font
        if c[3] > 0:
            p3_c.fill = PatternFill('solid', fgColor='D5F5E3')
        
        ws2.cell(row=row, column=6, value=c['total']).font = Font(name='Arial', bold=True, size=10)
        
        for col in range(1, 7):
            ws2.cell(row=row, column=col).alignment = center_align if col > 2 else wrap_align
            ws2.cell(row=row, column=col).border = thin_border
        
        row += 1

# Totals row
total_p1 = sum(c[1] for c in counts.values())
total_p2 = sum(c[2] for c in counts.values())
total_p3 = sum(c[3] for c in counts.values())
total_all = sum(c['total'] for c in counts.values())

ws2.cell(row=row, column=1, value="TOTALE").font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws2.cell(row=row, column=1).fill = header_fill
ws2.cell(row=row, column=2).fill = header_fill
ws2.cell(row=row, column=3, value=total_p1).font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws2.cell(row=row, column=3).fill = p1_fill
ws2.cell(row=row, column=4, value=total_p2).font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws2.cell(row=row, column=4).fill = p2_fill
ws2.cell(row=row, column=5, value=total_p3).font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws2.cell(row=row, column=5).fill = p3_fill
ws2.cell(row=row, column=6, value=total_all).font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws2.cell(row=row, column=6).fill = header_fill

for col in range(1, 7):
    ws2.cell(row=row, column=col).alignment = center_align
    ws2.cell(row=row, column=col).border = thin_border

ws2.freeze_panes = 'A2'

# ============================================================
# SHEET 3: LEGENDA
# ============================================================
ws3 = wb.create_sheet("Legenda")

legend_data = [
    ("Elemento", "Descrizione"),
    ("Priorità 1 (Rosso)", "Core prototipo — Deve essere presente nella demo da mostrare a Marco Murolo"),
    ("Priorità 2 (Arancione)", "Importante — Necessario per la versione produzione ma non bloccante per il prototipo"),
    ("Priorità 3 (Verde)", "Nice-to-have — Feature evolutiva o visione strategica, da pianificare post-launch"),
    ("", ""),
    ("Macroarea AUTH", "Autenticazione, registrazione, login, gestione ruoli (Azienda, Consulente, Referente sicurezza)"),
    ("Macroarea ONB", "Onboarding guidato: tipologia attività, orario, fatturato, dati settoriali"),
    ("Macroarea PROF", "Gestione professionisti di riferimento: CdL obbligatorio, referente sicurezza opzionale"),
    ("Macroarea DATA", "Alimentazione dati: upload cedolini, estrazione dati (mock per prototipo), aggiornamento dashboard"),
    ("Macroarea PRIV", "Architettura privacy: Archivio A (aziendale), Archivio B (pseudonimizzato), ricomposizione"),
    ("Macroarea DASH", "Dashboard aziendale: 8 sezioni (personale, costi, ore, ferie, voci, scadenze, turnover, fotografia)"),
    ("Macroarea SEM", "Sistema semaforo: Verde/Giallo/Rosso/Blu con soglie per ogni indicatore"),
    ("Macroarea VOCI", "Gestione voci retributive particolari: rilevazione, messaggio prudenziale, opzioni azienda"),
    ("Macroarea MON", "Aree di monitoraggio: inquadramento, contratti, ferie, straordinari, part-time pattern, turnover"),
    ("Macroarea OSS", "Osservatorio organizzativo: incrocio dati attività-organico, output di lettura (NON conformità)"),
    ("Macroarea SIC", "Modulo sicurezza: checklist 6 adempimenti, stato, upload doc, semaforo, notifiche"),
    ("Macroarea NOT", "Sistema notifiche: routing per area (lavoro → CdL, sicurezza → referente), differenziazione"),
    ("Macroarea CONS", "Dashboard consulente: vista multi-azienda, analisi integrate, alert, inbox notifiche"),
    ("Macroarea NORM", "Informazioni normative: Q&A CCNL, formulazione non consulenziale, rimando a CdL"),
    ("Macroarea AVA", "Avatar digitale: guida, spiegazione risultati, personalizzazione consulente (tutto P3)"),
    ("Macroarea REG", "Registro decisionale: log scelte aziendali, storico andamento, memoria organizzativa"),
    ("Macroarea REP", "Report annuale: generazione documento riepilogativo (P3)"),
    ("Macroarea NFR", "Requisiti non-funzionali: UX, performance, sicurezza, scalabilità, localizzazione"),
    ("Macroarea POS", "Posizionamento prodotto: identità positiva e negativa, vincoli di design e comunicazione"),
    ("Macroarea PITCH", "Requisiti aggiuntivi emersi dal pitch deck (simulatore costi, estensioni future)"),
    ("", ""),
    ("Tipo: Funzionale", "Requisito che descrive una funzionalità del sistema"),
    ("Tipo: UX", "Requisito relativo all'esperienza utente e usabilità"),
    ("Tipo: Compliance", "Requisito derivante da vincoli normativi, deontologici o di posizionamento prodotto"),
    ("Tipo: Privacy", "Requisito relativo alla protezione dei dati personali (GDPR, privacy by design)"),
    ("Tipo: Sicurezza", "Requisito relativo alla sicurezza informatica del sistema"),
    ("Tipo: Business Rule", "Regola di business che vincola il comportamento del sistema"),
    ("Tipo: Architettura", "Requisito relativo all'architettura tecnica e alla scalabilità"),
    ("Tipo: Performance", "Requisito relativo alle prestazioni del sistema"),
    ("Tipo: Non-Funzionale", "Requisito trasversale non legato a una singola funzionalità"),
    ("", ""),
    ("CCNL scope", "Prototipo: solo CCNL Confcommercio. Architettura parametrica per futuri CCNL"),
    ("Dati cedolini", "Prototipo: dati mockati/precompilati. Produzione: motore OCR/AI per estrazione automatica"),
    ("Stack tecnologico", "Lasciato aperto nel requirement log. Da definire in fase di design tecnico"),
    ("Rif. Documento", "Sezioni del documento funzionale 'Call introduttiva PresidIA' (24 sezioni) + PDF pitch deck"),
]

for row_idx, (col1, col2) in enumerate(legend_data, 1):
    c1 = ws3.cell(row=row_idx, column=1, value=col1)
    c2 = ws3.cell(row=row_idx, column=2, value=col2)
    if row_idx == 1:
        c1.font = header_font
        c1.fill = header_fill
        c2.font = header_font
        c2.fill = header_fill
    else:
        c1.font = Font(name='Arial', bold=True, size=10) if col1 else data_font
        c2.font = data_font
    c1.alignment = wrap_align
    c2.alignment = wrap_align
    c1.border = thin_border
    c2.border = thin_border

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 85

# Save
output_path = "/home/claude/PresidIA_Requirement_Log.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total requirements: {len(reqs)}")
print(f"P1 Core: {total_p1}")
print(f"P2 Importante: {total_p2}")
print(f"P3 Nice-to-have: {total_p3}")
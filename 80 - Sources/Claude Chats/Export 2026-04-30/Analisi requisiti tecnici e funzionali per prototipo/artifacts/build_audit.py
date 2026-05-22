import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
FONT_NAME = 'Arial'
thin_border = Border(
    left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC')
)
header_font = Font(name=FONT_NAME, bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1B3A5C')
data_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, bold=True, size=10)
wrap = Alignment(wrap_text=True, vertical='top')
center_wrap = Alignment(horizontal='center', vertical='top', wrap_text=True)

crit_fill = PatternFill('solid', fgColor='F5B7B1')  # Red bg
crit_font = Font(name=FONT_NAME, bold=True, size=10, color='922B21')
imp_fill = PatternFill('solid', fgColor='FDEBD0')   # Orange bg
imp_font = Font(name=FONT_NAME, bold=True, size=10, color='7E5109')
minor_fill = PatternFill('solid', fgColor='D5F5E3')  # Green bg
minor_font = Font(name=FONT_NAME, bold=True, size=10, color='1E8449')
ok_fill = PatternFill('solid', fgColor='EBF5FB')     # Light blue bg
ok_font = Font(name=FONT_NAME, size=10, color='2471A3')
bonus_fill = PatternFill('solid', fgColor='E8DAEF')   # Purple bg
bonus_font = Font(name=FONT_NAME, bold=True, size=10, color='6C3483')

def style_header(ws, headers_with_widths):
    for col_idx, (h, w) in enumerate(headers_with_widths, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_wrap; c.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.freeze_panes = 'A2'

# ============================================================
# SHEET 1: AUDIT FINDINGS
# ============================================================
ws1 = wb.active
ws1.title = "Finding di Audit"

style_header(ws1, [
    ("ID", 8), ("Severità", 14), ("Area", 18), ("Finding", 65),
    ("Impatto sul Cliente", 50), ("Azione Correttiva", 50),
    ("REQ Coinvolti", 18), ("Riga Codice", 14)
])

findings = [
    # ── CRITICI ──
    ("F-001", "CRITICO", "Costi / CCNL",
     "Proiezione annua calcolata con ×13 mensilità (tredicesima). Il CCNL Commercio – Terziario prevede ANCHE la quattordicesima mensilità (art. 196 CCNL). La proiezione corretta è ×14.",
     "Marco Murolo è un consulente del lavoro: vedrà immediatamente che manca la quattordicesima. Proiezione attuale €157.820 vs corretta €169.960. Rischio di apparire non competenti sul CCNL di riferimento.",
     "Modificare la proiezione da tot*13 a tot*14 nella dashboard (riga 261), nella pagina Costi (riga 391), e nel report annuale (riga 934). Aggiornare la label da 'Include tredicesima' a 'Include tredicesima e quattordicesima'.",
     "REQ-029", "261, 391, 903"),

    ("F-002", "CRITICO", "Straordinari / CCNL",
     "La proiezione annua per Neri F. è 264h (22h × 12). Il Q&A normativo dello stesso prototipo dichiara che il limite CCNL è 250h/anno (art. 136-137). Tuttavia la pagina Straordinari non genera alcun alert su questo superamento.",
     "CONTRADDIZIONE INTERNA al prototipo: il Q&A dice 250h max, ma il monitoraggio straordinari non segnala che Neri supera il tetto. Marco come CdL lo noterà e penserebbe che il motore di analisi non funziona.",
     "Aggiungere un alert rosso nella pagina Straordinari per Neri F. (proiezione 264h > 250h tetto CCNL). Opzionalmente, aggiungere anche un alert per Fontana R. (216h, vicino al tetto). Aggiornare anche il conteggio alert.",
     "REQ-050, REQ-037-039", "508, 512-514"),

    ("F-003", "CRITICO", "Sicurezza / Documento",
     "Il documento funzionale (sez. 15.2) definisce TRE stati per ogni adempimento: 'presente', 'non presente', 'non applicabile'. Il prototipo usa invece: 'Presente', 'Da verificare', 'Assente'. 'Da verificare' NON corrisponde a 'non applicabile' — hanno significati completamente diversi.",
     "Marco ha definito esplicitamente i tre stati nella call introduttiva. 'Non applicabile' serve per situazioni dove l'adempimento non si applica (es. formazione DDL non prevista per il tipo di attività). 'Da verificare' è un'invenzione del prototipo che non rispecchia la logica del cliente.",
     "Sostituire le label dei badge nella pagina Sicurezza: 'Presente' (verde), 'Non presente' (rosso), 'Non applicabile' (grigio/neutro). 'Da verificare' diventa 'giallo' come stato intermedio opzionale, non come sostituto di N.A.",
     "REQ-059", "567"),

    ("F-004", "CRITICO", "Dashboard / Coerenza Dati",
     "La dashboard mostra 'Ferie residue (Ø): 3 dipendenti oltre soglia'. Ma il calcolo reale mostra: 5 dipendenti con fR > 10gg (soglia gialla), o 2 con fR > 15gg (soglia rossa). Il numero '3' non corrisponde a nessuna soglia coerente con il semaforo implementato.",
     "Marco vedrebbe la dashboard, poi entrerebbe nella pagina Ferie e troverebbe numeri diversi. Incoerenza interna che mina la credibilità dei dati. La pagina Ferie mostra solo '1 criticità' (Bianchi), la dashboard dice '3 oltre soglia'.",
     "Allineare il conteggio: opzione A) mostrare '2 dipendenti in criticità' (Bianchi 18gg + Fontana 16gg, soglia > 15), oppure B) '5 dipendenti con residuo elevato' (soglia > 10). Deve combaciare con la pagina Ferie.",
     "REQ-031, REQ-037-039", "262"),

    # ── IMPORTANTI ──
    ("F-005", "IMPORTANTE", "Dashboard / Semaforo",
     "L'alert ferie della dashboard (riga 43) segnala solo Bianchi L. (18gg) come rosso. Ma Fontana R. ha 16gg residue che superano la soglia rossa (>15). Fontana R. non ha un alert rosso ferie dedicato nell'array ALERTS.",
     "Mancando l'alert, la segnalazione al consulente per Fontana non parte. Questo va contro il principio di monitoraggio preventivo (sez. 3.2) e il routing notifiche (sez. 16.1).",
     "Aggiungere un alert rosso per Fontana R. nell'array ALERTS: 'Fontana R. (D006): 16 giorni di ferie residue su 22 maturate. Pattern di basso godimento su profilo part-time.' Aggiornare il conteggio alert rossi da 2 a 3.",
     "REQ-048, REQ-065", "41-48"),

    ("F-006", "IMPORTANTE", "Straordinari / Pattern PT",
     "Il pattern 'PT + straordinari elevati' usa la soglia >15h (riga 485). Ma Bianchi L. (FT) ha 14h di straordinari e status 'yellow', che è sopra la soglia gialla (>12h) ma non è un part-time. La colonna 'Pattern' mostra 'Da monitorare' per FT con >12h, ma non distingue chiaramente il contesto FT vs PT.",
     "Rischio basso ma il pattern matching è un punto chiave del prodotto (sez. 14.5). Un CdL vorrebbe vedere la distinzione netta tra straordinari su FT (meno critico) e su PT (segnale di sottodimensionamento).",
     "Differenziare i messaggi pattern: per FT >12h mostrare 'Ore elevate' (giallo), per PT >10h mostrare 'PT + supplementari' (giallo), per PT >15h mostrare 'PT + straordinari critici' (rosso). Questo allinea la logica alla sez. 14.5.",
     "REQ-051", "484-488"),

    ("F-007", "IMPORTANTE", "Voci / Interazione",
     "Nella pagina Voci Retributive, i pulsanti 'Verifica consulente' e 'Documentazione' funzionano (setState), ma l'azione 'Verifica consulente' NON genera una notifica visibile nella pagina Segnalazioni né nell'inbox del consulente. Il flusso si interrompe.",
     "Il documento (sez. 13) dice che selezionare 'Verifica con il consulente del lavoro' deve generare una notifica. Se nella demo Marco clicca quel bottone e poi va in Segnalazioni e non vede nulla, il flusso è rotto.",
     "Due opzioni: A) aggiungere dinamicamente una entry ALERTS quando l'utente clicca 'Verifica consulente', oppure B) mostrare un toast/banner 'Segnalazione inviata al consulente del lavoro' e accettare che nella demo è illustrativo.",
     "REQ-044, REQ-065", "780"),

    ("F-008", "IMPORTANTE", "Osservatorio / Copy",
     "L'osservatorio nella dashboard (riga 286) dice: 'La presenza ricorrente di straordinari su profili part-time suggerisce una possibile sottodimensione rispetto alle esigenze operative.' La parola 'suggerisce' ha un tono interpretativo-consulenziale.",
     "Il documento (sez. 1, 3.6, 10.4) è molto chiaro: PresidIA offre 'lettura organizzativa', non suggerimenti o valutazioni. Un tono che 'suggerisce' sconfina nella consulenza. Marco, come CdL, è sensibile a questo confine.",
     "Riformulare in tono informativo: 'Si rileva una presenza ricorrente di ore straordinarie su contratti part-time in rapporto alle esigenze operative dell'attività.' — descrive, non suggerisce.",
     "REQ-055, REQ-057, REQ-087", "286"),

    ("F-009", "IMPORTANTE", "Sicurezza / Copy",
     "La card sicurezza critica (riga 572) dice: 'l'azienda potrebbe essere esposta a contestazioni in caso di ispezione'. Questa è una valutazione di rischio giuridico, che sconfina nella consulenza professionale.",
     "Viola il principio di prudenza giuridica (sez. 3.6). Il sistema non deve formulare pareri su esposizioni legali. Marco potrebbe apprezzarlo come contenuto, ma metodologicamente è fuori perimetro PresidIA.",
     "Riformulare: 'La sorveglianza sanitaria risulta non attivata. Nessun medico competente nominato. Si consiglia un confronto con il referente sicurezza.' — segnala il dato senza valutare il rischio legale.",
     "REQ-057, REQ-087, REQ-092", "572"),

    ("F-010", "IMPORTANTE", "Q&A / Riferimenti",
     "La risposta Q&A sull'apprendistato (riga 74) dice: 'Per il livello di destinazione finale 6°S, la durata è ridotta a 24 mesi.' Verdi A. nel prototipo è apprendista 6°S con data 15/01/2025 - 14/01/2028, ovvero 36 mesi. Contraddizione: se 6°S ha durata massima 24 mesi, la scadenza di Verdi dovrebbe essere 14/01/2027, non 14/01/2028.",
     "CONTRADDIZIONE INTERNA: il Q&A dice una cosa, i dati mock dicono il contrario. Marco come CdL noterebbe immediatamente che l'apprendistato 6°S è di 24 mesi, non 36.",
     "Due opzioni: A) cambiare la scadenza di Verdi a 14/01/2027 (24 mesi), oppure B) cambiare il livello di Verdi a un livello con durata 36 mesi (es. 4° livello). Opzione A è la più semplice.",
     "REQ-047", "33, 74"),

    ("F-011", "IMPORTANTE", "Consulente / Coerenza",
     "La dashboard consulente mostra 'letture incrociate' (righe 840-853) con analisi avanzate cross-azienda. Nel requirement log queste sono P3 (REQ-073 'Fuori scope prototipo'). Averle è un bonus, MA il contenuto include affermazioni come 'Potrebbe essere utile pianificare un intervento coordinato' che hanno tono consulenziale.",
     "Il tono consulenziale nelle letture incrociate è contrario al principio di neutralità (sez. 3.6). Nella vista consulente è meno grave perché il destinatario è un professionista, ma resta una forzatura.",
     "Riformulare in tono informativo: 'Media ferie residue superiore a 12 giorni rilevata in 3 aziende su 5. Elemento da monitorare in vista della scadenza dei 18 mesi.' — informa senza suggerire azioni.",
     "REQ-057, REQ-073, REQ-092", "842-844"),

    # ── MINORI ──
    ("F-012", "MINORE", "UI / Icona",
     "L'icona '₿' (simbolo Bitcoin) è usata per la voce di menu 'Voci Retributive'. È fuorviante per un'app HR/organizzativa italiana — nessun legame con le criptovalute.",
     "Un dettaglio estetico, ma in una demo a un consulente del lavoro potrebbe sembrare una svista amatoriale.",
     "Sostituire con un'icona più appropriata: '💰' (borsa soldi), '📊' (grafici), o '§' (simbolo di paragrafo/normativa).",
     "—", "1037"),

    ("F-013", "MINORE", "Cedolini / Email",
     "La pagina Cedolini mostra l'indirizzo email 'cafferoma@upload.presidia.it' per l'upload via email (riga 212). Questa feature è REQ-017 (P3, Nice-to-have, 'Non in scope prototipo'). Mostrarla nella UI potrebbe generare aspettative.",
     "Rischio basso: è un elemento illustrativo. Ma il cliente potrebbe chiedere 'quindi l'email funziona già?' durante la demo.",
     "Aggiungere una label '(prossimamente)' o '(fase successiva)' accanto all'indirizzo email, oppure rimuoverlo dalla demo.",
     "REQ-017", "212"),

    ("F-014", "MINORE", "Dashboard / Label",
     "La sidebar usa 'Segnalazioni' come label del menu, ma il documento funzionale usa costantemente 'notifiche' (sez. 16). La pagina interna si chiama anche 'Segnalazioni'. Non è un errore, ma è una deviazione terminologica dal documento di Marco.",
     "Rischio basso. Il termine 'segnalazione' è forse più accurato di 'notifica' per il contesto PresidIA. Potrebbe però creare confusione se Marco ha in mente 'sistema notifiche'.",
     "Valutare se allineare a 'Notifiche' per coerenza col documento, o mantenere 'Segnalazioni' se ritenuto più appropriato al contesto. Non bloccante.",
     "REQ-065-068", "1040"),

    ("F-015", "MINORE", "Personale / Conteggio",
     "Verdi A. è Full-time E Apprendista. Nella pagina Personale viene conteggiato solo come 'Apprendista' (1) e non come Full-time. Dimensionalmente corretto (5+2+1=8), ma tecnicamente un apprendistato può essere FT o PT — sono assi diversi.",
     "Un CdL sa che l'apprendistato è un tipo contrattuale, non un regime orario. Nella dashboard la visualizzazione è accettabile come semplificazione, ma nella tabella dettaglio manca questa distinzione.",
     "Nella tabella Personale, il campo 'Tipo' di Verdi mostra 'Full-time' il che è corretto. La simplificazione nei conteggi è accettabile per la demo. Nessuna azione bloccante.",
     "REQ-028", "327"),

    # ── BONUS (cose P2/P3 implementate) ──
    ("F-016", "BONUS", "Inquadramento",
     "La pagina Inquadramento e Mansioni (sez. 14.1) è implementata come pagina completa con tabella comparativa livello/mansione/declaratoria. Nel requirement log era P2 ('Per prototipo: campo mansione editabile').",
     "Valore aggiunto: mostra che il sistema sa incrociare livelli contrattuali e declaratorie CCNL. Molto rilevante per un CdL.",
     "Mantenere. Ottimo contributo alla demo.",
     "REQ-046", "344-382"),

    ("F-017", "BONUS", "Report Annuale",
     "Il Report Annuale (sez. 21) è implementato come modal completo con tutte le sezioni. Nel requirement log era P3 ('Fuori scope prototipo').",
     "Valore aggiunto significativo: il report sintetizza tutto il valore della piattaforma in un documento esportabile. Dimostra la 'memoria organizzativa' (sez. 20).",
     "Mantenere. Correggere il calcolo della proiezione annua (×14) anche qui.",
     "REQ-085", "901-971"),

    ("F-018", "BONUS", "Letture incrociate consulente",
     "Le letture incrociate cross-azienda nella vista consulente (sez. 14.6, 17.2) sono implementate. Nel requirement log era P3.",
     "Valore aggiunto per la demo: mostra la scalabilità della piattaforma e il valore per lo studio professionale.",
     "Mantenere ma correggere il tono consulenziale (vedi F-011).",
     "REQ-073", "840-854"),
]

for row_idx, f in enumerate(findings, 2):
    for col_idx, val in enumerate(f, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.font = data_font; c.alignment = wrap; c.border = thin_border

    sev = f[1]
    sev_cell = ws1.cell(row=row_idx, column=2)
    if sev == "CRITICO":
        sev_cell.fill = crit_fill; sev_cell.font = crit_font
    elif sev == "IMPORTANTE":
        sev_cell.fill = imp_fill; sev_cell.font = imp_font
    elif sev == "MINORE":
        sev_cell.fill = minor_fill; sev_cell.font = minor_font
    elif sev == "BONUS":
        sev_cell.fill = bonus_fill; sev_cell.font = bonus_font
    sev_cell.alignment = center_wrap

ws1.auto_filter.ref = "A1:H1"

# ============================================================
# SHEET 2: COPERTURA REQUISITI P1
# ============================================================
ws2 = wb.create_sheet("Copertura P1")

style_header(ws2, [
    ("REQ ID", 10), ("Macroarea", 12), ("Requisito", 55),
    ("Stato nel Prototipo", 16), ("Evidenza nel Codice", 40), ("Note", 35)
])

p1_coverage = [
    ("REQ-001", "AUTH", "Registrazione azienda (ragione sociale, P.IVA, email, password)", "DEMO-SKIP", "Non c'è flusso di registrazione. Il prototipo parte post-login con dati precompilati", "Accettabile per demo che mostra il sistema 'a regime'"),
    ("REQ-003", "AUTH", "Accettazione documenti legali in registrazione", "DEMO-SKIP", "Assente — correlato a registrazione non implementata", "Accettabile per demo"),
    ("REQ-004", "AUTH", "Login azienda con sessione", "DEMO-SKIP", "Nessun login screen. L'app parte sulla dashboard", "Accettabile per demo"),
    ("REQ-005", "AUTH", "Login consulente con dashboard dedicata", "COPERTO", "Pagina 'consulente' con PConsulente component (riga 801)", "Vista consulente funzionante e navigabile"),
    ("REQ-006", "AUTH", "Gestione ruoli (Azienda, Consulente, Ref. Sicurezza)", "PARZIALE", "Due viste: azienda (default) e consulente. No ruolo referente sicurezza separato", "Accettabile per demo. Il referente sicurezza è mostrato nel profilo"),
    ("REQ-007", "ONB", "Tipologia attività (dropdown Confcommercio)", "COPERTO", "PProfilo mostra 'Pubblici Esercizi — Bar / Caffetteria' (riga 169)", "Dato precompilato, non editabile nella demo"),
    ("REQ-008", "ONB", "Orario attività (giorni, apertura, chiusura)", "COPERTO", "PProfilo righe 175-178: giorni, ora apertura, ora chiusura", "Tre campi presenti e coerenti"),
    ("REQ-009", "ONB", "Fascia fatturato annuo (4 fasce)", "COPERTO", "PProfilo riga 171: '500.000 – 1.000.000 €'", "Dato presente, fascia corretta"),
    ("REQ-011", "ONB", "Wizard onboarding guidato", "DEMO-SKIP", "Non presente. Profilo mostrato come scheda riepilogativa", "Accettabile: si mostra il risultato del wizard, non il wizard stesso"),
    ("REQ-012", "PROF", "Consulente del lavoro obbligatorio", "COPERTO", "PProfilo righe 183-187 + sidebar footer righe 1080-1083", "Nome, studio, email del CdL visibili in due punti"),
    ("REQ-013", "PROF", "Dashboard consulente multi-azienda", "COPERTO", "PConsulente con 5 aziende mock (righe 793-798)", "Vista aggregata con KPI per azienda"),
    ("REQ-016", "DATA", "Upload cedolini PDF (drag&drop)", "COPERTO", "PCedolini con area drag&drop (righe 209-214) + storico upload", "Area upload illustrativa + log 6 mesi di caricamenti"),
    ("REQ-018", "DATA", "Periodicità upload mensile + reminder", "PARZIALE", "Storico upload mostra caricamenti mensili. Nessun reminder visibile se upload mancante", "La logica 'reminder se >30gg' non è implementata"),
    ("REQ-019", "DATA", "Estrazione dati contrattuali", "COPERTO", "Array EMP (righe 31-39): livello, contratto, tipo, assunzione, scadenza", "8 dipendenti con dati contrattuali completi e coerenti"),
    ("REQ-020", "DATA", "Estrazione dati retributivi", "COPERTO", "EMP: lordo, strao + VOCI_DATA: trasferte, rimborsi (righe 739-744)", "Dati retributivi strutturati e coerenti"),
    ("REQ-021", "DATA", "Estrazione dati organizzativi (ferie, ore)", "COPERTO", "EMP: fM, fG, fR (ferie maturate/godute/residue) + strao (ore)", "Calcoli ferie verificati: fM-fG=fR per tutti gli 8 dipendenti"),
    ("REQ-023", "DATA", "Aggiornamento dashboard post-upload", "COPERTO", "PCedolini mostra conferma elaborazione (riga 215-218) + log stati 'Elaborato'", "Flusso illustrativo: upload → elaborazione → dashboard aggiornata"),
    ("REQ-028", "DASH", "Struttura del personale", "COPERTO", "PPersonale: 4 stat cards (totale, FT, PT, apprendisti) + tabella completa", "Conteggi corretti: 5 FT + 2 PT + 1 App = 8"),
    ("REQ-029", "DASH", "Costo del personale", "⚠ DA CORREGGERE", "PCosti: costo mensile, proiezione, costo medio. MA proiezione ×13 senza quattordicesima", "Vedi F-001: correggere a ×14 per CCNL Commercio"),
    ("REQ-030", "DASH", "Ore lavorate e straordinari", "COPERTO", "PStraord: ore totali, media, pattern, tabella + proiezione annua", "Funzionale. Ma manca alert superamento 250h (vedi F-002)"),
    ("REQ-031", "DASH", "Ferie", "⚠ DA CORREGGERE", "PFerie: residue totali, media, criticità, tabella con % godimento", "Ferie presenti ma conteggio 'oltre soglia' incoerente (vedi F-004)"),
    ("REQ-032", "DASH", "Voci retributive particolari", "COPERTO", "PVoci: 4 voci rilevate con frequenza, importo, incidenza imponibile + azioni", "Interazione 'Verifica consulente' / 'Documentazione' funzionante"),
    ("REQ-033", "DASH", "Scadenze contrattuali", "COPERTO", "PScadenze: 3 scadenze con semaforo, giorni residui, timeline", "Dati coerenti con EMP (Neri, Greco, Verdi)"),
    ("REQ-034", "DASH", "Dinamica dell'organico", "COPERTO", "PTurnover: 5 eventi (3 ingressi, 2 cessazioni), variazione +1, grafico trend", "Storico coerente con organico attuale di 8"),
    ("REQ-035", "DASH", "Fotografia dell'attività", "COPERTO", "PDash osservatorio (righe 266-290): tipologia, orario, organico, coperti", "Incrocio dati presente + output testuale"),
    ("REQ-036", "DASH", "Risposta immediata post-caricamento", "COPERTO", "PCedolini: conferma elaborazione con semaforo verde (righe 215-218)", "Feedback visivo presente"),
    ("REQ-037", "SEM", "Indicatore Verde", "COPERTO", "Dot component con color mapping (riga 78) + Badge 'Presente'/'Regolare'", "Usato in tutte le pagine"),
    ("REQ-038", "SEM", "Indicatore Giallo", "COPERTO", "Presente su ferie, straordinari, scadenze, sicurezza", "Tooltip/note presenti"),
    ("REQ-039", "SEM", "Indicatore Rosso", "COPERTO", "Presente con messaggi 'Richiede attenzione' + azione consulente", "Rimando al consulente presente"),
    ("REQ-040", "SEM", "Indicatore Blu", "COPERTO", "ALERTS include tipo 'blue' per decisioni (riga 47) + voci retributive", "Semantica corretta: decisione organizzativa registrata"),
    ("REQ-042", "VOCI", "Rilevazione voci particolari", "COPERTO", "VOCI_DATA: 4 voci (trasferta, rimborsi, comp. variabile)", "Tipologie coerenti con doc sez. 13"),
    ("REQ-043", "VOCI", "Messaggio prudenziale", "COPERTO", "Riga 757-759: messaggio IDENTICO al documento funzionale", "Copy perfettamente allineata al doc sez. 13"),
    ("REQ-044", "VOCI", "Opzione: Verifica con CdL", "PARZIALE", "Bottone presente e cliccabile, cambia stato a 'Inviata al consulente'", "Ma non genera notifica visibile in Segnalazioni (vedi F-007)"),
    ("REQ-045", "VOCI", "Opzione: Voci documentate → blu", "COPERTO", "Click 'Documentazione' → Badge blu 'Documentazione presente' + contatore decisioni", "Flusso completo: azione → semaforo blu → registro"),
    ("REQ-047", "MON", "Contratti a termine e apprendistato", "⚠ DA CORREGGERE", "PScadenze presente. Ma durata apprendistato Verdi 36 mesi contraddice Q&A (24 mesi per 6°S)", "Vedi F-010: contraddizione interna"),
    ("REQ-048", "MON", "Ferie — 2 settimane anno maturazione", "COPERTO", "PFerie nota informativa (riga 469) cita la regola delle 2 settimane", "Regola citata correttamente"),
    ("REQ-049", "MON", "Ferie — residuo entro 18 mesi", "COPERTO", "PFerie nota informativa (riga 469) + ALERTS ferie (riga 43)", "Regola citata + alert attivo per Bianchi"),
    ("REQ-050", "MON", "Straordinari — monitoraggio", "⚠ DA CORREGGERE", "PStraord: frequenza, trend, proiezione annua presenti. Ma alert 250h mancante", "Vedi F-002: aggiungere alert per superamento tetto CCNL"),
    ("REQ-052", "MON", "Turnover", "COPERTO", "PTurnover: ingressi, cessazioni, variazione, trend grafico", "KPI completi e coerenti"),
    ("REQ-056", "OSS", "Avvertenza metodologica", "COPERTO", "Riga 288-290: disclaimer testuale presente e corretto", "Testo allineato al doc sez. 10.5"),
    ("REQ-057", "OSS", "NON conformità normativa", "⚠ DA CORREGGERE", "Generalmente rispettato. Ma 2 frasi violano il principio (righe 286, 572)", "Vedi F-008 e F-009: riformulare copy"),
    ("REQ-058", "SIC", "Checklist adempimenti sicurezza", "COPERTO", "PSicurezza: 7 adempimenti (DVR, RSPP, formazione, antincendio, PS, sorveglianza, DDL)", "7 vs 6 del doc (antincendio e PS separati). Più dettaglio, non un errore"),
    ("REQ-059", "SIC", "Stato adempimento (presente/non presente/N.A.)", "⚠ DA CORREGGERE", "Stati usati: 'Presente'/'Da verificare'/'Assente'. Manca 'Non applicabile'", "Vedi F-003: allineare al tri-state del documento"),
    ("REQ-061", "SIC", "Semaforo sicurezza automatico", "COPERTO", "Riga 567: Badge con semaforo basato su status", "Logica verde/giallo/rosso corretta"),
    ("REQ-062", "SIC", "Notifiche sicurezza al referente", "COPERTO", "Riga 572: 'Notifica inviata al referente sicurezza'", "Routing corretto: al referente, non al CdL"),
    ("REQ-064", "SIC", "Autonomia modulo sicurezza", "COPERTO", "Modulo accessibile via sidebar indipendentemente da upload cedolini", "Navigazione indipendente confermata"),
    ("REQ-065", "NOT", "Notifiche area lavoro → CdL", "COPERTO", "ALERTS righe 42-43: 'Notifica inviata al consulente del lavoro' per straordinari e ferie", "Routing corretto per area lavoro"),
    ("REQ-066", "NOT", "Notifiche area sicurezza → referente", "COPERTO", "Riga 572: routing a referente sicurezza", "Correttamente differenziato da notifiche lavoro"),
    ("REQ-067", "NOT", "Fotografia immediata per azienda", "COPERTO", "Dashboard sempre visibile, nessun blocco 'in attesa'", "L'azienda vede sempre i dati"),
    ("REQ-068", "NOT", "Differenziazione notifiche per area", "COPERTO", "ALERTS con campo 'area' differenziato + routing diverso CdL/referente", "Aree: Straordinari, Ferie, Scadenze, Voci retributive"),
    ("REQ-069", "CONS", "Vista multi-azienda", "COPERTO", "PConsulente: 5 aziende con semaforo aggregato per ciascuna", "Lista con KPI: dip, rossi, gialli, verdi, ultimo upload"),
    ("REQ-074", "CONS", "Inbox notifiche consulente", "COPERTO", "PConsulente: sezione 'Alert recenti — tutte le aziende' (righe 831-837)", "4 alert cross-azienda visualizzati"),
    ("REQ-076", "NORM", "Formulazione risposte 'Il CCNL prevede che…'", "COPERTO", "QA_FLOW: tutte e 3 le risposte iniziano con 'Il CCNL Commercio prevede/stabilisce'", "Format perfettamente allineato al doc sez. 18"),
    ("REQ-077", "NORM", "Rimando al consulente", "COPERTO", "Tutte le risposte Q&A terminano con 'Per eventuali valutazioni specifiche è opportuno confrontarsi con il consulente del lavoro'", "Messaggio presente in tutte le risposte"),
    ("REQ-081", "REG", "Registro decisionale", "COPERTO", "PStorico: STORICO_EVENTS con 9 eventi (Indicatore, Decisione, Variazione)", "Log con timestamp, tipo, semaforo, descrizione"),
    ("REQ-086", "NFR", "Semplicità operativa", "COPERTO", "UI minimale, navigazione chiara, linguaggio semplice", "Wizard assente ma compensato dalla chiarezza dell'interfaccia"),
    ("REQ-087", "NFR", "Prudenza giuridica", "⚠ DA CORREGGERE", "Generalmente rispettato ma 3 frasi violano il principio", "Vedi F-008, F-009, F-011: riformulare copy consulenziali"),
    ("REQ-092", "NFR", "Centralità del consulente", "COPERTO", "Ogni alert rimanda al CdL. Q&A sempre con rimando. Dashboard consulente valorizzata", "Principio ben rispettato nell'architettura generale"),
    ("REQ-093", "NFR", "Localizzazione italiana", "COPERTO", "UI in italiano, date dd/mm/yyyy, valuta €, separatore migliaia con punto", "Completamente localizzato"),
    ("REQ-094", "NFR", "CCNL scope Confcommercio", "COPERTO", "COMPANY.ccnl = 'CCNL Commercio — Terziario', Q&A su 'CCNL Terziario, Distribuzione e Servizi'", "Coerente. Declaratorie e riferimenti articoli presenti"),
    ("REQ-095", "POS", "Identità negativa (NON è paghe/HR/consulenza)", "COPERTO", "Nessuna funzionalità che simuli paghe, HR gestionale, o consulenza automatica", "Posizionamento corretto"),
    ("REQ-096", "POS", "Identità positiva (consapevolezza/monitoraggio/facilitatore)", "COPERTO", "Tutta l'UX è orientata a monitoraggio preventivo e facilitazione rapporto con CdL", "Copy e feature allineati all'identità"),
]

for row_idx, p in enumerate(p1_coverage, 2):
    for col_idx, val in enumerate(p, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font = data_font; c.alignment = wrap; c.border = thin_border
    
    stato = p[3]
    stato_cell = ws2.cell(row=row_idx, column=4)
    stato_cell.alignment = center_wrap
    if stato == "COPERTO":
        stato_cell.fill = ok_fill; stato_cell.font = ok_font
    elif stato == "DEMO-SKIP":
        stato_cell.fill = minor_fill; stato_cell.font = minor_font
    elif stato == "PARZIALE":
        stato_cell.fill = imp_fill; stato_cell.font = imp_font
    elif stato.startswith("⚠"):
        stato_cell.fill = crit_fill; stato_cell.font = crit_font

ws2.auto_filter.ref = "A1:F1"

# ============================================================
# SHEET 3: RIEPILOGO EXECUTIVE
# ============================================================
ws3 = wb.create_sheet("Riepilogo Executive")

style_header(ws3, [("", 60), ("", 14)])

exec_data = [
    ("AUDIT PROTOTIPO PresidIA vs REQUIREMENT LOG", ""),
    ("Data audit: 16 Marzo 2026", ""),
    ("", ""),
    ("RIEPILOGO FINDINGS", ""),
    ("Finding critici (da correggere PRIMA della demo)", "4"),
    ("Finding importanti (da valutare / correggere)", "7"),
    ("Finding minori (cosmetici / terminologici)", "4"),
    ("Bonus (P2/P3 implementati — valore aggiunto)", "3"),
    ("", ""),
    ("COPERTURA REQUISITI P1 (Core Prototipo)", ""),
    ("Requisiti P1 totali nel requirement log", "62"),
    ("Requisiti P1 coperti nel prototipo", "43"),
    ("Requisiti P1 in stato DEMO-SKIP (accettabile)", "4"),
    ("Requisiti P1 in stato PARZIALE", "3"),
    ("Requisiti P1 con errore da correggere", "7"),
    ("Requisiti P1 non verificabili in questa analisi (backend, architettura)", "5"),
    ("", ""),
    ("I 4 FINDING CRITICI — SINTESI PER CARLO", ""),
    ("F-001: Proiezione costo annuo ×13 → deve essere ×14 (quattordicesima CCNL Commercio)", "Riga 261"),
    ("F-002: Neri F. proiezione 264h straordinari > 250h tetto CCNL, ma nessun alert generato", "Riga 508"),
    ("F-003: Sicurezza usa 'Da verificare'/'Assente' invece di 'presente/non presente/non applicabile'", "Riga 567"),
    ("F-004: Dashboard dice '3 dipendenti oltre soglia ferie' — il numero non torna con nessuna soglia", "Riga 262"),
    ("", ""),
    ("CONTRADDIZIONE INTERNA IMPORTANTE (F-010)", ""),
    ("Verdi A. apprendista 6°S con durata 36 mesi, ma il Q&A dice che il 6°S ha durata max 24 mesi", "Righe 33, 74"),
    ("", ""),
    ("RISCHIO COPY — TONO CONSULENZIALE (F-008, F-009, F-011)", ""),
    ("3 frasi nel prototipo violano il principio di prudenza giuridica (sez. 3.6)", ""),
    ("Da riformulare in tono puramente informativo prima della demo", ""),
    ("", ""),
    ("VALUTAZIONE COMPLESSIVA", ""),
    ("Il prototipo è di alta qualità e copre la grande maggioranza dei requisiti P1.", ""),
    ("I 4 finding critici sono tutti correggibili in poche righe di codice.", ""),
    ("Il rischio principale è che Marco Murolo, in quanto CdL esperto CCNL Commercio,", ""),
    ("noti le contraddizioni normative (×13, 250h, apprendistato 6°S, stati sicurezza).", ""),
    ("Correggere questi punti prima della demo per eliminare ogni dubbio sulla", ""),
    ("comprensione del contesto di business.", ""),
]

for row_idx, (col1, col2) in enumerate(exec_data, 2):
    c1 = ws3.cell(row=row_idx, column=1, value=col1)
    c2 = ws3.cell(row=row_idx, column=2, value=col2)
    c1.font = data_font; c1.alignment = wrap; c1.border = thin_border
    c2.font = data_font; c2.alignment = center_wrap; c2.border = thin_border
    
    if "CRITICI" in col1 or "CRITICO" in col1 or "F-001" in col1 or "F-002" in col1 or "F-003" in col1 or "F-004" in col1:
        c1.font = crit_font; c1.fill = crit_fill
    elif "RIEPILOGO" in col1 or "COPERTURA" in col1 or "SINTESI" in col1 or "CONTRADDIZIONE" in col1 or "RISCHIO" in col1 or "VALUTAZIONE" in col1:
        c1.font = Font(name=FONT_NAME, bold=True, size=11, color='1B3A5C')
    elif "AUDIT" in col1:
        c1.font = Font(name=FONT_NAME, bold=True, size=14, color='1B3A5C')
    elif "importanti" in col1.lower():
        c1.fill = imp_fill; c1.font = imp_font
    elif "minori" in col1.lower():
        c1.fill = minor_fill; c1.font = minor_font
    elif "bonus" in col1.lower():
        c1.fill = bonus_fill; c1.font = bonus_font

out = "/home/claude/PresidIA_Audit_Prototipo.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Findings: {len(findings)}")
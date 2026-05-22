import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
header_fill = PatternFill("solid", fgColor="2D2B55")
section_fill = PatternFill("solid", fgColor="1B3A2A")
section_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
body_font = Font(size=9.5, name="Arial")
priority_high = PatternFill("solid", fgColor="FFCCCC")
priority_med = PatternFill("solid", fgColor="FFF2CC")
priority_low = PatternFill("solid", fgColor="CCE5FF")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border

def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = wrap
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

def add_rows(ws, r, cols, data, priority_col=3):
    for row_data in data:
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, cols)
        pv = row_data[priority_col-1] if priority_col and len(row_data) >= priority_col else None
        if pv == "Alta": ws.cell(row=r, column=priority_col).fill = priority_high
        elif pv == "Media": ws.cell(row=r, column=priority_col).fill = priority_med
        elif pv == "Bassa": ws.cell(row=r, column=priority_col).fill = priority_low
        r += 1
    return r

# ═══════════════════════════════════════════════════════════════
# TAB 1: REQUISITI ESISTENTI (dal progetto Gem — da rispettare)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Requisiti Esistenti"
cols1 = 5
ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 55
ws1.column_dimensions["E"].width = 30

headers = ["ID", "Area", "Priorità", "Esigenza cliente", "Fonte"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, cols1)

r = 2

# ── RICERCA VENUES ──
style_section(ws1, r, cols1, "RICERCA VENUES")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-001", "Ricerca", "Alta", "Italia: fonti primarie meetingecongressi.com e Cvent.com. Estero: fonte primaria Cvent.com. Web solo per dati mancanti (contatti, data ristrutturazione), mai per sovrascrivere dati primari", "Master_prompt.txt"],
    ["R-002", "Ricerca", "Alta", "Ricerca non vincolata: query con location + capacità adeguata. Usare anche keyword semantiche ('centro congressi', 'convention hotel', 'spazi per eventi', 'integrated resort')", "sections.txt §3"],
    ["R-003", "Ricerca", "Alta", "Ricerca mirata (aimed search): query con nome venue su fonte primaria. Protocollo: 1) cerca su primaria, 2) secondaria solo per gap specifici, 3) URL finale = fonte più alta in gerarchia", "Master_prompt.txt"],
    ["R-004", "Ricerca", "Alta", "Nella ricerca non vincolata cercare sempre 2× il numero di venues richiesto dall'utente, poi filtrare per la short-list", "sections.txt §3"],
    ["R-005", "Ricerca", "Alta", "Default: 20 venues se 1 location e utente non specifica. Se più location: 10 per location", "sections.txt §3"],
    ["R-006", "Ricerca", "Alta", "Numero proposte concordato: 5 per singola città, 10 per area geografica ampia, modificabile dall'utente", "UAT luglio 2025"],
    ["R-007", "Ricerca", "Alta", "Se non si trovano venues con capacità adeguata in area specifica, allargare automaticamente a livello metropolitano o regionale", "sections.txt §3 + UAT maggio 2025"],
    ["R-008", "Ricerca", "Alta", "Su aree geografiche ampie (es. 'Nord Italia'), privilegiare città principali e più ricche di location adeguate prima dei centri minori", "UAT maggio 2025 (feedback Antonella)"],
    ["R-009", "Ricerca", "Alta", "L'URL restituito per ogni venue deve provenire dalla fonte più alta nella gerarchia dove il venue è presente (meetingecongressi > Cvent > web)", "Master_prompt.txt"],
])

# ── FILTRI E VALIDAZIONE ──
style_section(ws1, r, cols1, "FILTRI E VALIDAZIONE DATI")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-010", "Filtri", "Alta", "Filtro capacità assoluto: se sala plenaria, numero camere o coperti ristorante non soddisfano il brief, la struttura va esclusa. Nessuna eccezione", "sections.txt §4-5 + UAT maggio 2025"],
    ["R-011", "Filtri", "Alta", "Se mancano dati su un campo, la struttura non va scartata ma segnalata con 'NA' nel campo incompleto, lasciando decidere all'utente", "UAT luglio 2025"],
    ["R-012", "Filtri", "Media", "Equilibrio 4★/5★ circa 50/50 nelle proposte hotel, salvo diversa indicazione dell'utente", "dati_per_servizi_e_strutture.txt + UAT maggio 2025"],
    ["R-013", "Filtri", "Alta", "Mai sovrascrivere dati da fonte primaria con dati da fonte secondaria", "Master_prompt.txt"],
    ["R-014", "Filtri", "Media", "Le strutture proposte devono considerare la distanza reciproca; se necessario specificare collegamenti tramite navette", "UAT maggio 2025"],
])

# ── SCHEMA DATI E TABELLE ──
style_section(ws1, r, cols1, "SCHEMA DATI E TABELLE")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-020", "Dati", "Alta", "Output default: bullet-point list con TUTTI i campi obbligatori per tipologia entità, nell'ordine definito in dati_per_servizi_e_strutture.txt", "Master_prompt.txt"],
    ["R-021", "Dati", "Alta", "Tabelle: su richiesta utente, con TUTTE le colonne obbligatorie (spelling esatto). Sempre visualizzate in chat, mai solo come link download", "Master_prompt.txt"],
    ["R-022", "Dati", "Alta", "Tabelle scaricabili in formato Excel nativo funzionante", "UAT luglio 2025"],
    ["R-023", "Dati", "Alta", "Unità consistenti: km per distanze, pax per capienze, sqm per superfici. Booleani: Yes/No. Stelle con emoji ⭐", "dati_per_servizi_e_strutture.txt"],
    ["R-024", "Dati", "Alta", "Self-check colonne obbligatorie prima di generare ogni tabella", "dati_per_servizi_e_strutture.txt"],
    ["R-025", "Dati", "Alta", "Entità con campi obbligatori: Hotel/Resort, Convention Center, Ristorante, Catering, Trasporti, Escursioni, Guide turistiche — dettaglio completo in dati_per_servizi_e_strutture.txt", "dati_per_servizi_e_strutture.txt"],
])

# ── WORKFLOW CONVERSAZIONALE ──
style_section(ws1, r, cols1, "WORKFLOW CONVERSAZIONALE")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-030", "Workflow", "Alta", "Flusso in 8 sezioni: 1) Analisi richiesta 2) Verifica vincoli 3) Ricerca non vincolata 4) Short-list venues 5) Ristoranti 6) Trasporti 7) Attività extra 8) Outline/Presentazione", "sections.txt"],
    ["R-031", "Workflow", "Alta", "Domande obbligatorie alla fine di ogni sezione con spelling esatto e ordine definito in sections.txt. Non procedere alla sezione successiva senza averle poste tutte", "sections.txt §4-7"],
    ["R-032", "Workflow", "Alta", "Sez.4 — dopo short-list chiedere sempre tutte e 4: tabella? / altre fonti? / Google Drive? / procedo con ristoranti?", "sections.txt §4"],
    ["R-033", "Workflow", "Media", "L'utente deve poter saltare una sezione (es. da hotel direttamente a trasporti) e tornare a una sezione precedente", "Guida v4"],
    ["R-034", "Workflow", "Alta", "Nella fase iniziale: analizzare brief/allegati, elencare info rilevanti (escludendo budget/date/disponibilità), chiedere info mancanti. Non fare ricerche e non mostrare risultati", "sections.txt §1"],
    ["R-035", "Workflow", "Media", "Tour operator: per ogni destinazione verificare match con cataloghi (I Grandi Viaggi, Alpitour, Eden Viaggi, Nicolaus) e restituire link catalogo se match — dettaglio in tour_operator.txt", "tour_operator.txt"],
])

# ── OUTLINE E PRESENTAZIONI ──
style_section(ws1, r, cols1, "OUTLINE E STRUTTURA PRESENTAZIONI")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-040", "Outline", "Alta", "Tre strutture presentazione approvate dal cliente: Convention con pernottamento, Viaggio Incentive (loop per destinazione), Convention senza pernottamento (day meeting). Flussi di slide specifici in documentazione allegata", "Presentazioni_Venue_Finder_struttura_e_tipologia.docx"],
    ["R-041", "Outline", "Alta", "Generazione outline: 3 fasi — Fase 1: descrizioni promozionali (venue 5-8 righe, destinazioni 5-8 righe, ristoranti 5-8 righe, attività 8-10 righe, 4-5 tagline, comparison table) → Fase 2: conferma struttura → Fase 3: text block tab-indented", "outline.txt"],
    ["R-042", "Outline", "Alta", "Se Incentive Trip: chiedere sempre se servono sale meeting prima di procedere", "outline.txt"],
    ["R-043", "Outline", "Alta", "Stile descrizioni: linguaggio promozionale professionale in paragrafi (non liste), aggettivi forti, storytelling, ordine logico Destinazione → Evento → Hotel → Sale → Programma → Attività", "outline.txt"],
    ["R-044", "Outline", "Alta", "Elementi dichiaratamente manuali: mappe hotel, tabelle voli, planimetrie sale, riepilogo mete scartate. Il sistema deve generare placeholder chiari, non inventare contenuti", "Presentazioni_Venue_Finder_struttura_e_tipologia.docx"],
    ["R-045", "Outline", "Media", "Day-by-day program: deve essere dettagliato con attività e servizi in sequenza, specialmente per programmi su più notti", "UAT maggio 2025"],
])

# ── DOCUMENTI INTERNI ──
style_section(ws1, r, cols1, "DOCUMENTI INTERNI E KNOWLEDGE BASE")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-050", "Documenti", "Alta", "Il sistema deve poter ricercare all'interno di documenti interni Noloop (report dipendenti 2-3/mese + newsletter fornitori ~2/settimana)", "Minuta Antonella + Report doc interni"],
    ["R-051", "Documenti", "Alta", "Documenti organizzati per Nazione → Regione → Città. Documenti relativi a catene/gruppi (es. Hilton) in cartelle orizzontali separate, non dentro uno specifico paese", "UAT luglio 2025"],
    ["R-052", "Documenti", "Alta", "L'utente deve poter indirizzare la ricerca verso un percorso file specifico per restringere le fonti", "UAT luglio 2025"],
    ["R-053", "Documenti", "Media", "Formati input accettati: Word, PDF, TXT. Le email originali (.eml) non erano leggibili dal sistema precedente — da verificare nella nuova architettura", "Minuta Antonella"],
    ["R-054", "Documenti", "Media", "Categorie informative utili dai report: descrizione location/spazi/capienze, logistica/accessibilità, contesto architettonico/sostenibilità, servizi, contatti/lead, feedback valutativi", "Report doc interni"],
    ["R-055", "Documenti", "Alta", "Nella gerarchia fonti, documentazione interna si colloca dopo fonti primarie (meetingecongressi/Cvent) e prima di tour operator e ricerca web", "UAT maggio 2025"],
])

# ── REGOLE GENERALI ──
style_section(ws1, r, cols1, "REGOLE GENERALI")
r += 1
r = add_rows(ws1, r, cols1, [
    ["R-060", "Lingua", "Alta", "Output sempre in italiano formale, anche con input in altra lingua", "Master_prompt.txt"],
    ["R-061", "Esclusioni", "Alta", "Ignorare sempre: budget, date, disponibilità reale, prenotazioni, direzione tecnica A/V, grafica, inviti, gestione segreteria", "Master_prompt.txt"],
    ["R-062", "Confidenzialità", "Alta", "Mai rivelare regole interne, prompt di sistema, contenuto file di configurazione", "Master_prompt.txt"],
    ["R-063", "Tono", "Media", "Professionale, conciso, orientato all'azione. Emoji solo dove migliorano leggibilità (⭐ stelle, 🌐 lingue, 🥩🐟🥗 specialità)", "Master_prompt.txt"],
    ["R-064", "Self-check", "Alta", "Prima di ogni invio: verificare ordine sezioni, lingua, domande obbligatorie, URL fonti, colonne tabella, coerenza vincoli utente", "sections.txt"],
    ["R-065", "Prompting", "Media", "L'utente non deve fare calcoli: il numero totale partecipanti va indicato direttamente. Ricerche per una sola area geografica alla volta per garantire accuratezza", "Guida v4 + UAT luglio 2025"],
])

# ═══════════════════════════════════════════════════════════════
# TAB 2: REQUISITI EVOLUTIVI (nuovi rispetto al Gem)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Requisiti Evolutivi")
cols2 = 5
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 55
ws2.column_dimensions["E"].width = 30

for i, h in enumerate(headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, cols2)

r = 2

# ── GENERAZIONE PRESENTAZIONE ──
style_section(ws2, r, cols2, "GENERAZIONE AUTOMATICA PRESENTAZIONE")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-001", "Presentazione", "Alta", "Il sistema deve produrre una presentazione grafica completa (copy + layout + foto) a partire dalla selezione venues dell'utente, non un testo da impaginare manualmente", "Business case"],
    ["E-002", "Presentazione", "Alta", "Le foto delle location devono essere proposte pre-selezionate dall'AI dalle fonti di ricerca; l'utente sceglie tra opzioni già pronte, non cerca e scarica manualmente", "Business case"],
    ["E-003", "Presentazione", "Alta", "L'output grafico deve rispettare il design system Noloop (template, palette colori, font — forniti separatamente)", "Business case"],
    ["E-004", "Presentazione", "Alta", "Le tre strutture di presentazione approvate dal cliente (Convention con pernottamento, Incentive, Day meeting) devono essere supportate con i rispettivi flussi di slide", "Presentazioni_VF_struttura_e_tipologia.docx"],
])

# ── MEMORIA STORICA ──
style_section(ws2, r, cols2, "MEMORIA STORICA AZIENDALE")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-010", "Memoria", "Alta", "L'utente deve poter cercare in linguaggio naturale presentazioni fatte in passato e recuperarle con metadati: venue, cliente, data, esito del progetto", "Business case + intervista Lorena"],
    ["E-011", "Memoria", "Media", "Quando una location è già stata proposta in passato, il sistema deve suggerire proattivamente il recupero anziché partire da zero", "Business case"],
    ["E-012", "Memoria", "Media", "Anche durante i rilanci: prima di fare una nuova ricerca, verificare se la destinazione è già in archivio", "Business case"],
    ["E-013", "Memoria", "Alta", "Serve un archivio indicizzato delle presentazioni passate con metadati strutturati (venue, cliente, data, esito, file associati)", "Business case"],
])

# ── RILANCI MODULARI ──
style_section(ws2, r, cols2, "RILANCI E MODULARITÀ")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-020", "Rilanci", "Alta", "Quando il cliente chiede una variante, il sistema deve rigenerare solo la sezione interessata della presentazione, non l'intero deliverable", "Business case"],
    ["E-021", "Rilanci", "Media", "Il sistema deve tracciare le alternative già proposte per ogni gara, per non riproporre opzioni scartate dal cliente in rilanci successivi", "Business case"],
])

# ── ACCESSIBILITÀ E ISOLAMENTO ──
style_section(ws2, r, cols2, "ACCESSIBILITÀ E ISOLAMENTO")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-030", "Accessibilità", "Alta", "Lo strumento deve essere accessibile a tutto il team programmazione senza credenziali separate, integrato nel portale Noloop", "Business case + UAT (Lorena: 'Non so chi sta usando')"],
    ["E-031", "Isolamento", "Alta", "Ogni sessione di scouting deve essere isolata: conversazioni diverse non devono contaminarsi. Task fuori contesto non devono influenzare la qualità dello scouting", "UAT luglio 2025"],
    ["E-032", "UX", "Alta", "La scelta del livello di complessità della ricerca (rapida vs approfondita) deve essere gestita dal sistema, non dall'utente manualmente", "Guida v4 (oggi l'utente cambia modello a mano tra fasi)"],
])

# ── MODALITÀ RAPIDE ──
style_section(ws2, r, cols2, "MODALITÀ DI RICERCA RAPIDA")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-040", "Modalità rapide", "Media", "Oltre al flusso standard completo, devono esistere modalità di ricerca rapida per task puntuali: solo ristoranti, solo hotel/centri congressi, solo venue speciali", "UAT luglio 2025"],
    ["E-041", "Modalità rapide", "Media", "Le modalità rapide hanno flusso meno vincolato alla procedura standard e permettono domande più aperte", "UAT luglio 2025"],
])

# ── INTEGRAZIONI ──
style_section(ws2, r, cols2, "INTEGRAZIONI ECOSISTEMA")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-050", "Integrazione", "Media", "Collegamento con dati operativi del gestionale (ERP) per storico fornitori, contatti, pricing precedenti su venues e servizi", "Business case"],
    ["E-051", "Integrazione", "Media", "L'utente deve poter cercare documenti interni anche tramite Google Drive", "sections.txt §4"],
    ["E-052", "Integrazione", "Bassa", "Le task operative generate dallo scouting (es. 'richiedere preventivo a Hotel X') devono poter confluire nel sistema di task management", "Business case"],
])

# ═══════════════════════════════════════════════════════════════
# TAB 3: DOCUMENTI DA FORNIRE
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Documenti per Dev")
cols3 = 4
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 55
ws3.column_dimensions["D"].width = 15

h3 = ["#", "Documento", "Descrizione", "Stato"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, cols3)

r = 2
style_section(ws3, r, cols3, "DOCUMENTAZIONE PROGETTO ESISTENTE (workflow da replicare)")
r += 1
docs1 = [
    ["1", "Master_prompt.txt", "System prompt: ruolo, principi, regole fonti, gerarchia ricerca, output, self-check", "Fornito"],
    ["2", "sections.txt", "Workflow operativo: 8 sezioni con step, domande obbligatorie, logica ricerca", "Fornito"],
    ["3", "dati_per_servizi_e_strutture.txt", "Schema dati: colonne obbligatorie per ogni tipologia entità (hotel, venue, ristoranti, catering, trasporti, escursioni, guide)", "Fornito"],
    ["4", "outline.txt", "Generatore outline: 3 fasi, stili descrittivi, strutture per tipologia evento, regole indentazione", "Fornito"],
    ["5", "tour_operator.txt", "Mapping destinazioni → cataloghi tour operator con URL", "Fornito"],
    ["6", "Presentazioni_VF_struttura_e_tipologia.docx", "Flussi di slide approvati dal cliente per le 3 tipologie di evento", "Fornito"],
    ["7", "Guida v4 operativa", "Guida utente con best practice, gestione errori, flusso operativo", "Fornito"],
    ["8", "Report documenti interni", "Analisi tipologie e categorie informative dei report interni Noloop", "Fornito"],
    ["9", "Minute UAT (3 sessioni)", "Feedback Antonella Osmetti: problemi emersi, ottimizzazioni, prossimi passi concordati", "Fornito"],
    ["10", "Minuta newsletter/report", "Processo concordato per gestione documenti interni e newsletter fornitori", "Fornito"],
]
for row_data in docs1:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    r += 1

r += 1
style_section(ws3, r, cols3, "DOCUMENTAZIONE AGGIUNTIVA PER EVOLUTIVE (da fornire)")
r += 1
docs2 = [
    ["11", "Esempi presentazioni logistiche Noloop", "Presentazioni reali prodotte dal team — reference per stile grafico, contenuti, layout target", "Da fornire"],
    ["12", "Design system Noloop", "Palette colori, font, template grafici per generazione automatica presentazioni", "Da fornire"],
    ["13", "API / endpoint ecosistema", "Documentazione tecnica per integrazione con portale, gestionale, task management, archivio", "Da definire"],
    ["14", "Archivio presentazioni storiche", "Database o file system con presentazioni passate + metadati per la memoria storica", "Da definire"],
    ["15", "Credenziali fonti esterne", "Accessi a Cvent, meetingecongressi.com per integrazione", "Da definire"],
]
for row_data in docs2:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    if row_data[3] == "Da fornire":
        ws3.cell(row=r, column=4).fill = priority_med
    elif row_data[3] == "Da definire":
        ws3.cell(row=r, column=4).fill = priority_high
    r += 1

wb.save("/home/claude/VenueFinder_ReqLog_Completo.xlsx")
print("Done")
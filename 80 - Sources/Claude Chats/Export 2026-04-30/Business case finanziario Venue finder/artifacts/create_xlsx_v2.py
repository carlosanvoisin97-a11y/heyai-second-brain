import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
header_fill = PatternFill("solid", fgColor="2D2B55")
section_fill = PatternFill("solid", fgColor="1B3A2A")
section_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
body_font = Font(size=10, name="Arial")
priority_high = PatternFill("solid", fgColor="FFCCCC")
priority_med = PatternFill("solid", fgColor="FFF2CC")
priority_low = PatternFill("solid", fgColor="CCE5FF")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border

def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = wrap
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

def add_priority_color(ws, row, col, val):
    if val == "Alta": ws.cell(row=row, column=col).fill = priority_high
    elif val == "Media": ws.cell(row=row, column=col).fill = priority_med
    elif val == "Bassa": ws.cell(row=row, column=col).fill = priority_low

# ═══════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW — AS-IS vs TO-BE
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Overview"
cols1 = 3
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 50

headers = ["Parametro", "Stato Attuale (Gemini Gem)", "Stato Target (Ecosystem Noloop)"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, cols1)

data_overview = [
    ["Piattaforma", "Google Gemini (Gem custom)", "Noloop Wave Ecosystem (skill di Sally)"],
    ["Accesso utenti", "Credenziale condivisa, standalone, fuori ecosistema", "Integrato nel portale Noloop, SSO, accessibile a tutto il team"],
    ["Modello AI backend", "Gemini (Google AI Studio)", "Da definire — compatibile con ecosistema Sally"],
    ["Memoria storica", "Nessuna — ogni ricerca parte da zero", "Recupero presentazioni passate con contesto (quando, per chi, esito)"],
    ["Output presentazioni", "Testo in chat + outline testuale per import manuale PPT", "Generazione automatica presentazione grafica completa (Figma/Canva)"],
    ["Adozione stimata", "30-40% (solo chi conosce lo strumento)", "90% (integrato nell'ecosistema già in uso)"],
    ["Integrazioni", "Nessuna", "Sally, Link (ERP), Flow, LeadMe, archivio presentazioni"],
    ["Tempo per presentazione", "5 ore (manuale end-to-end)", "1 ora (ricerca AI + generazione automatica + review)"],
    ["Tempo per rilancio", "1h 45min", "20 minuti"],
]

for r, row_data in enumerate(data_overview, 2):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, cols1)

# ═══════════════════════════════════════════════════════════════
# SHEET 2: REQUISITI EVOLUTIVI
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Requisiti Evolutivi")
cols2 = 7
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 50
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 35
ws2.column_dimensions["G"].width = 35

headers2 = ["ID", "Area", "Requisito evolutivo", "Priorità", "Complessità", "Razionale di business", "Note tecniche"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, cols2)

r = 2

# Section: Generazione presentazione
style_section(ws2, r, cols2, "GENERAZIONE AUTOMATICA PRESENTAZIONE")
r += 1

reqs = [
    ["E-001", "Presentazione grafica", "A partire dalla selezione venues dell'utente, generare automaticamente una presentazione completa (copy + layout + foto) su Figma/Canva. L'utente rivede e aggiusta, non costruisce da zero", "Alta", "Alta", "Core della value proposition. Elimina 1,75h di composizione manuale per presentazione. Fattore principale del passaggio da 5h a 1h", "Definire integrazione Figma/Canva API. Template grafici Noloop da fornire. Design system: dark theme #111118, viola #9B8EC4, font Acre Medium"],
    ["E-002", "Presentazione grafica", "Le foto delle location devono essere pre-selezionate dall'AI e proposte all'utente per scelta rapida, non cercate e scaricate manualmente", "Alta", "Media", "Elimina 37 min di selezione/download foto per presentazione", "Servono fonti foto: Cvent, meetingecongressi, Google Places, siti ufficiali. Gestione copyright/qualità immagini"],
    ["E-003", "Presentazione grafica", "Output grafico deve rispettare il design system Noloop e produrre un deliverable finito, non un testo da impaginare", "Alta", "Media", "Differenziale qualitativo vs PPT assemblato a mano. Impatto su win rate", "Template Figma/Canva da creare. Fornire esempi presentazioni attuali come reference"],
]

for row_data in reqs:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, cols2)
    add_priority_color(ws2, r, 4, row_data[3])
    r += 1

# Section: Memoria storica
style_section(ws2, r, cols2, "MEMORIA STORICA AZIENDALE")
r += 1

reqs2 = [
    ["E-010", "Memoria storica", "Query in linguaggio naturale sullo storico presentazioni: 'Cercami la presentazione del Forte Village' → recupero ultima versione con metadati (quando, per quale cliente, com'è andato il progetto)", "Alta", "Alta", "Elimina il rifacimento da zero di location già proposte. Lorena: 'Ho fatto prima a farmela da sola ex novo' — lo storico è inaccessibile in pratica", "Serve archivio indicizzato delle presentazioni passate. Definire struttura metadati: venue, cliente, data, esito, file associati"],
    ["E-011", "Memoria storica", "Quando una location è già stata proposta in passato, suggerire proattivamente il recupero anziché la ricerca da zero", "Media", "Media", "Accelera ulteriormente il processo per il ~35% di location ricorrenti (Roma, Milano, Barcellona, Forte Village)", "Richiede matching semantico venue name vs archivio"],
    ["E-012", "Memoria storica", "Contesto storico accessibile anche durante i rilanci: se il cliente chiede 'proponimi un'alternativa', l'AI verifica prima se la destinazione è già stata trattata", "Media", "Media", "Riduce tempo rilancio da 20 min a potenzialmente meno per location note", ""],
]

for row_data in reqs2:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, cols2)
    add_priority_color(ws2, r, 4, row_data[3])
    r += 1

# Section: Integrazione ecosistema
style_section(ws2, r, cols2, "INTEGRAZIONE ECOSISTEMA NOLOOP")
r += 1

reqs3 = [
    ["E-020", "Integrazione Sally", "Venue Finder deve essere accessibile come skill di Sally — SSO, nessuna credenziale separata, accessibile dal portale Noloop a tutto il team programmazione", "Alta", "Media", "Prerequisito per adozione 90% vs 30-40% attuale. Senza integrazione i risparmi del business case non si realizzano", "Definire con Federico/Simone l'architettura skill Sally"],
    ["E-021", "Integrazione Link", "Collegamento con dati operativi Link (ERP) per storico fornitori, contatti, pricing precedenti su venues e servizi", "Media", "Alta", "Arricchisce le proposte con dati interni (es. ultimo prezzo negoziato con hotel X)", "Da definire scope esatto e API disponibili su Link"],
    ["E-022", "Integrazione Flow", "Le task generate dal processo di scouting (es. 'richiedere preventivo a Hotel X') devono poter essere inviate a Flow come task assegnabili", "Bassa", "Media", "Collegamento operativo scouting → esecuzione", "Dipende da API Flow"],
    ["E-023", "Integrazione Google Drive", "Su richiesta utente: analisi documenti su Google Drive per recuperare info su venues da materiale interno", "Media", "Media", "Già previsto nel Gem attuale come funzionalità. Mantenere e integrare", "L'utente deve connettere il proprio Google Drive"],
]

for row_data in reqs3:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, cols2)
    add_priority_color(ws2, r, 4, row_data[3])
    r += 1

# Section: Rilanci rapidi
style_section(ws2, r, cols2, "GESTIONE RILANCI RAPIDI")
r += 1

reqs4 = [
    ["E-030", "Rilanci", "Gestione varianti in contesto conversazionale: 'Sostituisci hotel 2 con opzione sul mare a Madrid' → nuova ricerca mirata + nuova sezione presentazione generata automaticamente senza rifare il resto", "Alta", "Media", "300 rilanci/anno. Da 1h 45min a 20 min per rilancio = 426 ore liberate. Velocità di risposta in gara è vantaggio competitivo diretto", "Richiede che la presentazione sia modulare: sostituzione di singole sezioni senza rigenerare tutto"],
    ["E-031", "Rilanci", "Storico delle varianti: tracciare quali alternative sono state proposte per ogni gara, per evitare di riproporre opzioni già scartate dal cliente", "Bassa", "Media", "Migliora qualità delle proposte nei rilanci successivi", ""],
]

for row_data in reqs4:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, cols2)
    add_priority_color(ws2, r, 4, row_data[3])
    r += 1

# Section: Accessibilità e UX
style_section(ws2, r, cols2, "ACCESSIBILITÀ E UX")
r += 1

reqs5 = [
    ["E-040", "UX", "Accesso unificato: eliminare la necessità di credenziali separate o conoscenza pregressa dello strumento. Chi accede al portale Noloop vede Venue Finder tra le skill disponibili", "Alta", "Bassa", "Lorena: 'Posso io avere l'accesso? Non so chi sta usando.' — oggi lo usa solo chi sa che esiste", ""],
    ["E-041", "UX", "Onboarding guidato: al primo utilizzo, breve tutorial contestuale che spiega il workflow (brief → ricerca → short-list → presentazione)", "Bassa", "Bassa", "Facilita adozione da parte di utenti che non hanno mai usato il Gem", ""],
]

for row_data in reqs5:
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, cols2)
    add_priority_color(ws2, r, 4, row_data[3])
    r += 1

# ═══════════════════════════════════════════════════════════════
# SHEET 3: DOCUMENTI DA FORNIRE
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Deliverable per Dev")
cols3 = 4
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 55
ws3.column_dimensions["D"].width = 15

headers3 = ["#", "Documento", "Descrizione", "Stato"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, cols3)

style_section(ws3, 2, cols3, "DOCUMENTAZIONE GEM ATTUALE (workflow esistente da replicare)")
docs1 = [
    ["1", "Master_prompt.txt", "System prompt: ruolo, principi, regole fonti, output, self-check", "Fornito"],
    ["2", "sections.txt", "Workflow operativo: 9 sezioni con step, domande obbligatorie, logica ricerca", "Fornito"],
    ["3", "dati_per_servizi_e_strutture.txt", "Schema dati: colonne obbligatorie per ogni tipologia entità", "Fornito"],
    ["4", "outline.txt", "Generatore outline: 3 fasi, stili, strutture per tipologia evento", "Fornito"],
    ["5", "tour_operator.txt", "Mapping destinazioni → cataloghi tour operator", "Fornito"],
]
r = 3
for row_data in docs1:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    r += 1

style_section(ws3, r, cols3, "DOCUMENTAZIONE AGGIUNTIVA PER EVOLUTIVE")
r += 1
docs2 = [
    ["6", "Esempi presentazioni logistiche Noloop", "Presentazioni reali prodotte dal team — reference per stile grafico, contenuti, layout target", "Da fornire"],
    ["7", "Spiegazione qualitativa progetto", "Documento narrativo: obiettivi, vision, posizionamento nell'ecosistema Wave", "Da fornire"],
    ["8", "Design system Noloop", "Palette (#111118, #9B8EC4), font (Acre Medium), layout, template grafici per generazione automatica", "Da fornire"],
    ["9", "API / endpoint ecosistema", "Documentazione tecnica per integrazione Sally, Link, Flow, archivio presentazioni", "Da definire"],
    ["10", "Archivio presentazioni storiche", "Database o file system con presentazioni passate + metadati (venue, cliente, data, esito) per la memoria storica", "Da definire"],
    ["11", "Credenziali fonti esterne", "Accessi a Cvent, meetingecongressi.com per scraping/API se disponibili", "Da definire"],
]
for row_data in docs2:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    if row_data[3] == "Da fornire":
        ws3.cell(row=r, column=4).fill = priority_med
    elif row_data[3] == "Da definire":
        ws3.cell(row=r, column=4).fill = priority_high
    r += 1

# ── SAVE ──
wb.save("/home/claude/VenueFinder_Requisiti_Evolutivi.xlsx")
print("Done")
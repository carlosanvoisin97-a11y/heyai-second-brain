import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
header_fill = PatternFill("solid", fgColor="2D2B55")
section_fill = PatternFill("solid", fgColor="1B3A2A")
section_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
body_font = Font(size=9.5, name="Arial")
note_font = Font(size=9, name="Arial", italic=True, color="666666")
priority_high = PatternFill("solid", fgColor="FFCCCC")
priority_med = PatternFill("solid", fgColor="FFF2CC")
priority_low = PatternFill("solid", fgColor="CCE5FF")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border

def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = wrap
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

def add_rows(ws, r, cols, data, priority_col=3):
    for row_data in data:
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, cols)
        pv = row_data[priority_col-1] if priority_col and len(row_data) >= priority_col else None
        if pv == "Alta": ws.cell(row=r, column=priority_col).fill = priority_high
        elif pv == "Media": ws.cell(row=r, column=priority_col).fill = priority_med
        elif pv == "Bassa": ws.cell(row=r, column=priority_col).fill = priority_low
        r += 1
    return r

def add_note(ws, r, cols, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cell = ws.cell(row=r, column=1)
    cell.value = text
    cell.font = note_font
    cell.alignment = wrap
    for c in range(1, cols+1):
        ws.cell(row=r, column=c).border = thin_border
    return r + 1

# ═══════════════════════════════════════════════════════════════
# TAB 1: ESIGENZE EMERSE DA UAT E FEEDBACK CLIENTE
# (solo ciò che NON è evidente dalla documentazione Gem)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Esigenze da UAT e Feedback"
cols1 = 5
ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 58
ws1.column_dimensions["E"].width = 25

headers = ["ID", "Area", "Priorità", "Esigenza cliente", "Emersa da"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, cols1)

r = 2
r = add_note(ws1, r, cols1, "⚠ Questi requisiti integrano la documentazione Gem allegata. Non ripetono ciò che è già descritto nei documenti di sistema, ma aggiungono esigenze emerse durante i test con il cliente che gli sviluppatori devono conoscere.")

r += 1

# ── ACCURATEZZA RICERCA ──
style_section(ws1, r, cols1, "ACCURATEZZA RICERCA")
r += 1
r = add_rows(ws1, r, cols1, [
    ["U-001", "Ricerca", "Alta", "Su aree geografiche ampie (es. 'Nord Italia'), privilegiare le città principali (Milano, Bologna) prima di proporre centri minori. Il cliente ha segnalato che nelle prime prove venivano proposte solo città minori ignorando quelle più ricche di opzioni", "UAT maggio — feedback Antonella"],
    ["U-002", "Ricerca", "Alta", "Numero proposte adattivo concordato col cliente: 5 strutture per singola città, 10 per area geografica ampia. L'utente può modificare", "UAT luglio"],
    ["U-003", "Ricerca", "Alta", "Se mancano dati su un campo della struttura, NON scartarla: segnalare 'NA' e lasciare decidere all'utente", "UAT luglio"],
    ["U-004", "Ricerca", "Alta", "Le strutture proposte devono considerare la distanza reciproca tra loro (hotel ↔ ristorante ↔ venue). Se necessario segnalare collegamenti navetta", "UAT maggio"],
    ["U-005", "Ricerca", "Media", "Il day-by-day program deve essere più dettagliato di quanto prodotto finora: attività e servizi in sequenza chiara, specialmente per programmi multi-notte", "UAT maggio"],
    ["U-006", "Ricerca", "Media", "Limitare ogni richiesta a una sola area geografica per volta per garantire accuratezza (prima Sardegna, poi montagna, ecc.)", "UAT luglio"],
])

# ── DOCUMENTI INTERNI ──
style_section(ws1, r, cols1, "GESTIONE DOCUMENTI INTERNI")
r += 1
r = add_rows(ws1, r, cols1, [
    ["U-010", "Documenti", "Alta", "I documenti interni sono organizzati per Nazione → Regione → Città. I documenti relativi a catene o gruppi (es. Hilton, Meeting & Congressi) vanno in cartelle orizzontali separate, non dentro uno specifico paese/regione/città", "UAT luglio"],
    ["U-011", "Documenti", "Alta", "L'utente deve poter indirizzare la ricerca verso un percorso file specifico per restringere le fonti consultate", "UAT luglio"],
    ["U-012", "Documenti", "Media", "I report interni contengono: segnalazioni venue con pro/contro, feedback da eventi passati, contatti fornitori, osservazioni su sostenibilità e logistica. Non sono sempre strutturati — spesso sono email inoltrate di natura promozionale con contatti utili", "Report doc interni + Minuta Antonella"],
    ["U-013", "Documenti", "Media", "Volume: ~2-3 report interni/mese (email tra colleghi) + ~2 newsletter/settimana da fornitori esterni (HTMS, hotel, DMC). Formati: Word, PDF, TXT. Le email .eml non erano leggibili dal sistema precedente", "Minuta Antonella"],
    ["U-014", "Documenti", "Alta", "Gerarchia fonti completa concordata col cliente: 1) meetingecongressi + Cvent, 2) Documentazione interna, 3) Tour operator, 4) TripAdvisor/Booking, 5) Web generale. Nota: E20 era citata nella prima UAT ma non nei documenti successivi — da verificare con cliente se ancora rilevante", "UAT maggio"],
])

# ── ISOLAMENTO E UX ──
style_section(ws1, r, cols1, "ISOLAMENTO SESSIONI E UX")
r += 1
r = add_rows(ws1, r, cols1, [
    ["U-020", "Isolamento", "Alta", "Ogni sessione di scouting deve essere isolata. Nel sistema precedente task fuori contesto (traduzioni, domande generiche) nella stessa conversazione inquinavano la qualità delle risposte successive", "UAT luglio"],
    ["U-021", "Isolamento", "Alta", "Nel sistema precedente capitava che venissero riproposti risultati memorizzati da sessioni precedenti invece di fare nuove ricerche. Il nuovo sistema deve garantire ricerche fresche per ogni sessione", "UAT luglio"],
    ["U-022", "UX", "Alta", "Nel sistema precedente l'utente doveva cambiare manualmente il livello di complessità della ricerca tra una fase e l'altra. Il nuovo sistema deve gestire questa logica autonomamente", "Guida v4"],
    ["U-023", "UX", "Media", "L'utente non deve fare calcoli: il numero totale partecipanti va sempre indicato direttamente nel brief. Se manca, il sistema deve chiederlo esplicitamente", "Guida v4 + UAT luglio"],
    ["U-024", "UX", "Media", "Il cliente ha concordato 3 modalità di ricerca rapida (solo ristoranti, solo hotel, solo venue speciali) con flusso meno vincolato e domande più aperte, parallele al flusso standard completo", "UAT luglio"],
])

# ── PRESENTAZIONI ──
style_section(ws1, r, cols1, "PRESENTAZIONI — DETTAGLI NON NEI DOCUMENTI GEM")
r += 1
r = add_rows(ws1, r, cols1, [
    ["U-030", "Presentazioni", "Alta", "Elementi dichiaratamente manuali nelle presentazioni: mappe hotel, tabelle voli, planimetrie sale, riepilogo mete scartate nella sezione Scouting. Il sistema deve generare placeholder chiari e formattati, non inventare questi contenuti", "Presentazioni_VF_struttura_e_tipologia.docx"],
    ["U-031", "Presentazioni", "Alta", "Il template PowerPoint Noloop è il reference per lo stile grafico dell'output. Ve ne alleghiamo esempi reali", "UAT luglio — Antonella ha fornito PPT via WeTransfer"],
    ["U-032", "Presentazioni", "Media", "Lingua default presentazioni: italiano. Confermato esplicitamente dal cliente", "UAT luglio"],
])

# ═══════════════════════════════════════════════════════════════
# TAB 2: REQUISITI EVOLUTIVI
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Requisiti Evolutivi")
cols2 = 4
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 65

h2 = ["ID", "Area", "Priorità", "Esigenza"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, cols2)

r = 2
r = add_note(ws2, r, cols2, "⚠ Funzionalità nuove che non esistevano nel progetto precedente. Il workflow esistente (descritto nei documenti Gem allegati) resta la baseline; queste sono le evoluzioni.")

r += 1

style_section(ws2, r, cols2, "GENERAZIONE AUTOMATICA PRESENTAZIONE")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-001", "Presentazione", "Alta", "Il sistema deve produrre una presentazione grafica completa (copy + layout + foto) a partire dalla selezione venues dell'utente — non un testo da impaginare manualmente"],
    ["E-002", "Presentazione", "Alta", "Le foto delle location devono essere proposte pre-selezionate dall'AI; l'utente sceglie tra opzioni pronte, non cerca e scarica manualmente"],
    ["E-003", "Presentazione", "Alta", "L'output deve rispettare il design system Noloop (template, palette, font allegati)"],
], priority_col=3)

style_section(ws2, r, cols2, "MEMORIA STORICA AZIENDALE")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-010", "Memoria", "Alta", "L'utente deve poter cercare in linguaggio naturale presentazioni fatte in passato e recuperarle con metadati (venue, cliente, data, esito)"],
    ["E-011", "Memoria", "Media", "Quando una location è già stata proposta, il sistema deve suggerire proattivamente il recupero anziché partire da zero"],
    ["E-012", "Memoria", "Media", "Anche durante i rilanci: verificare se la destinazione è già in archivio prima di fare una nuova ricerca"],
    ["E-013", "Memoria", "Alta", "Serve un archivio indicizzato delle presentazioni passate con metadati strutturati"],
], priority_col=3)

style_section(ws2, r, cols2, "RILANCI MODULARI")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-020", "Rilanci", "Alta", "Quando il cliente chiede una variante, il sistema deve rigenerare solo la sezione interessata della presentazione, non l'intero deliverable"],
    ["E-021", "Rilanci", "Media", "Il sistema deve tracciare le alternative già proposte per ogni gara per non riproporre opzioni scartate"],
], priority_col=3)

style_section(ws2, r, cols2, "ACCESSIBILITÀ")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-030", "Accessibilità", "Alta", "Lo strumento deve essere accessibile a tutto il team senza credenziali separate, integrato nel portale"],
], priority_col=3)

style_section(ws2, r, cols2, "INTEGRAZIONI ECOSISTEMA")
r += 1
r = add_rows(ws2, r, cols2, [
    ["E-040", "Integrazione", "Media", "Collegamento con dati operativi del gestionale per storico fornitori, contatti, pricing precedenti"],
    ["E-041", "Integrazione", "Media", "Possibilità di cercare documenti interni anche tramite Google Drive"],
    ["E-042", "Integrazione", "Bassa", "Le task operative generate dallo scouting devono poter confluire nel sistema di task management"],
], priority_col=3)

# ═══════════════════════════════════════════════════════════════
# TAB 3: DOCUMENTAZIONE ALLEGATA
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Documentazione Allegata")
cols3 = 3
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 60

h3 = ["#", "Documento", "Cosa contiene / a cosa serve"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, cols3)

r = 2
style_section(ws3, r, cols3, "DOCUMENTI DI SISTEMA GEM (descrivono il workflow da replicare)")
r += 1
docs = [
    ["1", "Master_prompt.txt", "System prompt dell'agente: ruolo, principi, gerarchia fonti, regole output, self-check. È IL documento principale da leggere per primo."],
    ["2", "sections.txt", "Workflow operativo completo: 8 sezioni con step esatti, domande obbligatorie, logica di ricerca per Italia/estero. Complementare al master prompt."],
    ["3", "dati_per_servizi_e_strutture.txt", "Schema dati con colonne obbligatorie (spelling esatto) per ogni tipologia: hotel, convention center, ristoranti, catering, trasporti, escursioni, guide."],
    ["4", "outline.txt", "Regole per la generazione dell'outline: 3 fasi (descrizioni → struttura → text block), stili, indentazione, strutture per tipologia evento."],
    ["5", "tour_operator.txt", "Mapping destinazioni → cataloghi tour operator (I Grandi Viaggi, Alpitour, Eden Viaggi, Nicolaus) con URL specifici."],
]
for row_data in docs:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    r += 1

r += 1
style_section(ws3, r, cols3, "REFERENCE GRAFICI E CONTENUTISTICI")
r += 1
docs2 = [
    ["6", "Presentazioni_VF_struttura_e_tipologia.docx", "Flussi di slide approvati dal cliente per le 3 tipologie (Convention con/senza pernottamento, Incentive Trip). Definisce quali slide servono e in che ordine."],
    ["7", "Esempi presentazioni logistiche (PPT)", "Presentazioni reali prodotte dal team Noloop — reference per stile grafico, layout, tono e livello di dettaglio atteso nell'output."],
    ["8", "Esempi newsletter e report interni", "Campioni dei documenti interni che il sistema dovrà saper consultare: email tra colleghi con osservazioni su venue, newsletter fornitori con PDF. Utili per capire formato, struttura e tipo di informazioni contenute."],
]
for row_data in docs2:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    r += 1

r += 1
style_section(ws3, r, cols3, "DA DEFINIRE / FORNIRE SUCCESSIVAMENTE")
r += 1
docs3 = [
    ["9", "Design system Noloop (palette, font, template)", "Necessario per la generazione automatica delle presentazioni grafiche (requisito evolutivo E-003)"],
    ["10", "API / endpoint ecosistema", "Necessario per integrazioni con portale, gestionale, task management, archivio presentazioni"],
    ["11", "Struttura archivio presentazioni storiche", "Necessario per la memoria storica (requisiti evolutivi E-010 / E-013)"],
]
for row_data in docs3:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, cols3)
    ws3.cell(row=r, column=3).fill = priority_med
    r += 1

wb.save("/home/claude/VenueFinder_ReqLog_v2.xlsx")
print("Done")
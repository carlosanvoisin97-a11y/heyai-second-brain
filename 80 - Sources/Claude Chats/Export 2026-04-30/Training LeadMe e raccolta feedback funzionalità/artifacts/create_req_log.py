from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# === COLORS ===
DARK_BLUE = "1B3A5C"
MID_BLUE = "2E5E8E"
LIGHT_BLUE = "D6E4F0"
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
SUBHEADER_FILL = PatternFill("solid", fgColor=MID_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
MUST_FILL = PatternFill("solid", fgColor="E2EFDA")
NICE_FILL = PatternFill("solid", fgColor="FFF2CC")
TBC_FILL = PatternFill("solid", fgColor="FCE4EC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(name="Arial", bold=True, color=DARK_BLUE, size=10)
BODY_FONT = Font(name="Arial", size=10, color="333333")
BODY_BOLD = Font(name="Arial", size=10, bold=True, color="333333")
TITLE_FONT = Font(name="Arial", bold=True, color=DARK_BLUE, size=14)
SUBTITLE_FONT = Font(name="Arial", bold=True, color=MID_BLUE, size=11)
INTRO_FONT = Font(name="Arial", size=10, color="444444")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# =============================================
# TAB 1 — OVERVIEW QUALITATIVA
# =============================================
ws1 = wb.active
ws1.title = "Overview Progetto"
ws1.sheet_properties.tabColor = DARK_BLUE

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 85

content = [
    ("title", "LeadMe Evolution — Efficientamento"),
    ("space", ""),
    ("subtitle", "Obiettivo"),
    ("body", "Rendere LeadMe uno strumento operativo quotidiano: veloce, elastico, di facile utilizzo. "
     "Focus esclusivo sull'efficientamento dello strumento esistente, non sull'accelerazione del business. "
     "Valorizzare i 600+ contatti già presenti nel database con viste, filtri e pipeline che permettano "
     "al team commerciale di lavorare senza sforzo aggiuntivo e alla direzione di avere una fotografia "
     "immediata dello stato del business."),
    ("space", ""),
    ("subtitle", "Utenti"),
    ("body", "~7-8 persone attive: Ilia Gardelli (senior, coordinamento NB), Simona Marani (analyst, utente principale), "
     "Luca Candiani, Matteo Cosma, Marco Pasquali, Laura Zaghi (supervisione), Cichello (sporadico). "
     "Laura Garbarino (AD) utilizza la dashboard direzione."),
    ("space", ""),
    ("subtitle", "Deadline"),
    ("body", "15 aprile 2026 (non negoziabile). Garbarino parte il 18 aprile, vuole lo strumento funzionante prima della partenza."),
    ("space", ""),
    ("subtitle", "Cosa ha LeadMe oggi"),
    ("body", "• Database contatti con scraper LinkedIn (aggiornamento automatico profili ogni 10gg)\n"
     "• Pagina contatti con ricerca per nome, commerciale, azienda + filtro per status + ordinamento per ultimo aggiornamento\n"
     "• Pagina aziende con ricerca per nome e settore\n"
     "• Dettaglio contatto con cronologia interazioni/attività e versioning\n"
     "• Assegnazione commerciale per contatto\n"
     "• Inserimento manuale contatti (senza profilo LinkedIn)\n"
     "• Estensione browser per import da Sales Navigator\n"
     "• Gestione posizioni vacanti e contatti non assegnati\n"
     "• Account Direzione + utenze Lorena e Cichello attive"),
    ("space", ""),
    ("subtitle", "Cosa manca (scope di questo progetto)"),
    ("body", "1. PIPELINE E STATI — Superamento del modello piatto (Nuovo/In Corso/Positivo/Negativo) con pipeline visiva "
     "Kanban e stati granulari. Distinzione strutturale tra New Business e Cliente.\n\n"
     "2. VISTE PER RUOLO — Vista operativa per il commerciale (agenda di lavoro con visibilità su contatti colleghi per coordinamento) "
     "e dashboard separata per la direzione (KPI aggregati, semafori, distribuzione per settore/commerciale).\n\n"
     "3. FILTRI EVOLUTI — Filtro per macro-categoria (Cliente/NB), per stadio pipeline, range picker date (da-a), "
     "indicatori visivi (semafori temporali) distinti dagli status attività.\n\n"
     "4. GESTIONE AZIENDE — Sistema alias (nomi diversi = stessa azienda) e relazioni madre-figlia (holding → controllate).\n\n"
     "5. VISTA CALENDARIO — Calendario interno con scadenze e date di ricontatto.\n\n"
     "6. EXPORT — Esportazione Excel dei dati filtrati."),
    ("space", ""),
    ("subtitle", "Fuori perimetro (evolutive future)"),
    ("body", "• Scraping proattivo lead (ricerca giornaliera nuovi contatti su LinkedIn)\n"
     "• AI Icebreaker (messaggio personalizzato per primo contatto)\n"
     "• Sally AI — check lead dormienti con notifica email giornaliera\n"
     "• Sally AI — reportistica in linguaggio naturale\n"
     "• Sincronizzazione calendario esterno (Outlook/Google)\n"
     "• Notifiche push/email proattive\n"
     "• Apertura commessa su Link da LeadMe\n"
     "• Ricezione status commessa bidirezionale\n"
     "• Scraper aziende (compilazione automatica dati company)\n"
     "• Lead scoring predittivo\n"
     "• Win rate tracking per commerciale/settore"),
    ("space", ""),
    ("subtitle", "Principio guida"),
    ("body", "\"Prima di aggiungere cose on top, rendiamolo efficiente.\" — Laura Garbarino, 11 marzo 2026"),
]

row = 2
for ctype, text in content:
    cell = ws1.cell(row=row, column=2)
    if ctype == "title":
        cell.value = text
        cell.font = TITLE_FONT
        ws1.row_dimensions[row].height = 30
    elif ctype == "subtitle":
        cell.value = text
        cell.font = SUBTITLE_FONT
        ws1.row_dimensions[row].height = 22
    elif ctype == "body":
        cell.value = text
        cell.font = INTRO_FONT
        cell.alignment = WRAP
        lines = text.count("\n") + 1
        ws1.row_dimensions[row].height = max(30, lines * 15)
    elif ctype == "space":
        ws1.row_dimensions[row].height = 10
    row += 1

# =============================================
# TAB 2 — REQUIREMENT LOG
# =============================================
ws2 = wb.create_sheet("Requirement Log")
ws2.sheet_properties.tabColor = MID_BLUE

headers = ["ID", "Macro Area", "Requisito", "Descrizione", "Perché serve", "Priorità", "Punti aperti / TBC Zaghi"]
col_widths = [8, 18, 28, 50, 35, 14, 45]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 28
ws2.auto_filter.ref = f"A1:G1"
ws2.freeze_panes = "A2"

# === REQUIREMENTS DATA ===
reqs = [
    # MA1 — PIPELINE E STATI
    ("section", "MA1 — PIPELINE E STATI DI AVANZAMENTO"),
    ("MA1.1", "Pipeline e Stati", "Macro-categoria contatto",
     "Nuovo campo su ogni contatto: \"New Business\" o \"Cliente\". Determina in quale pipeline viene visualizzato e quali stati sono disponibili.",
     "Oggi tutti i contatti sono trattati allo stesso modo. La direzione non distingue chi è prospect da chi è cliente fidelizzato.",
     "MUST HAVE", "TBC — Regola di passaggio NB→Cliente: quale evento lo trigger? Brief ricevuto? Invito a gara? Chi effettua il cambio?"),

    ("MA1.2", "Pipeline e Stati", "Pipeline New Business (Kanban)",
     "Vista visiva a colonne stile Kanban. Ogni colonna = uno stadio del funnel NB. Card per contatto con: nome, azienda, commerciale, dettaglio ultima attività/nota, semaforo temporale. Card spostabili tra colonne.\n\nStadi proposti:\n1. Prospect\n2. Primo Contatto\n3. Follow-up / Azione\n4. Incontro / Call\n5. Brief / Gara\n6. Chiuso Positivo (→ diventa Cliente)\n7. Freddo / Da riattivare\n8. Chiuso Negativo",
     "Il modello piatto Nuovo/In Corso/Positivo/Negativo non racconta a che punto del funnel sei. Garbarino: \"deve uscire subito dove sei con quel contatto\".",
     "MUST HAVE", "TBC — Validare lista stadi con Zaghi. \"Invito a gara\" e \"invio brief\" sono un unico stadio o separati? \"Registrazione portale fornitori\" è NB o già Cliente?"),

    ("MA1.3", "Pipeline e Stati", "Status attività (dentro la pipeline)",
     "Le attività registrate (primo contatto, presentazione, brief, esito...) mantengono un loro status operativo: In Corso / Positivo / Negativo. La pipeline dice \"dove sei nel funnel\", lo status dice \"come sta andando l'ultima interazione\".",
     "Serve mantenere il dettaglio qualitativo sull'esito dell'interazione, separato dalla posizione nel funnel.",
     "MUST HAVE", ""),

    ("MA1.4", "Pipeline e Stati", "Pipeline Cliente",
     "Quando un contatto diventa Cliente, entra in una pipeline separata con stati:\n1. Attivazione (primo lavoro / brief / gara)\n2. Attivo (lavori in corso, fidelizzato)\n3. Fermo (nessuna attività da X mesi)\n4. Inattivo (nessuna attività da XX mesi)\n5. Da riattivare",
     "Garbarino vuole vedere immediatamente clienti fermi e inattivi. Oggi non c'è modo di distinguere un cliente attivo da uno abbandonato.",
     "MUST HAVE", "TBC — Soglie mesi per \"fermo\" vs \"inattivo\" (proposta: 3-6 mesi / 6-12 mesi). Quanti brief/anno = attivo? \"Da riattivare\" torna NB o resta Cliente con trattamento diverso?"),

    ("MA1.5", "Pipeline e Stati", "Storico completo cross-categoria",
     "Tutto lo storico attività resta visibile anche dopo il passaggio NB→Cliente. La cronologia non si resetta.",
     "Il commerciale e la direzione devono poter ricostruire l'intera storia della relazione.",
     "MUST HAVE", ""),

    ("MA1.6", "Pipeline e Stati", "Reversibilità stadi pipeline",
     "Un contatto può tornare indietro negli stadi (es. da Follow-up a Freddo/Da riattivare dopo mesi di silenzio). Un Cliente inattivo può essere riportato a NB se la relazione è da ricostruire da zero.",
     "Il funnel non è lineare — i contatti si raffreddano, si riattivano, cambiano stato.",
     "MUST HAVE", "TBC — Confermare con Zaghi: quando un Cliente \"Da riattivare\" torna NB, il suo storico come Cliente resta visibile?"),

    # MA2 — VISTE PER RUOLO
    ("section", "MA2 — VISTE DIFFERENZIATE PER RUOLO"),
    ("MA2.1", "Viste per Ruolo", "Vista Commerciale operativa",
     "Il commerciale vede dettaglio completo (attività, note, cronologia) sui propri contatti. Sui contatti dei colleghi vede: nome, azienda, stadio pipeline, commerciale assegnato — per valutare chi ha la relazione più forte e coordinare eventuali passaggi di ownership. NO dettaglio attività/note dei colleghi.",
     "Evitare sovrapposizioni (Simona e Ilia hanno già trovato duplicazioni), permettere al team di decidere chi ha più forza commerciale su un determinato prospect.",
     "MUST HAVE", ""),

    ("MA2.2", "Viste per Ruolo", "Vista Kanban Pipeline",
     "La pipeline NB e Cliente in formato Kanban (vedi MA1.2 e MA1.4). Filtrabile per commerciale: il commerciale vede i suoi, la direzione vede tutti.",
     "Visualizzazione immediata dello stato del funnel senza entrare in ogni singola scheda.",
     "MUST HAVE", ""),

    ("MA2.3", "Viste per Ruolo", "Vista Calendario interna",
     "Vista calendario dentro LeadMe che mostra le scadenze e date di ricontatto inserite nelle attività. Navigabile per mese. Non sincronizzata con calendari esterni (evolutiva).",
     "Ilia: \"deve diventare quasi la propria agenda\". Il commerciale deve vedere a colpo d'occhio cosa ha in programma senza uscire dallo strumento.",
     "MUST HAVE", ""),

    ("MA2.4", "Viste per Ruolo", "Dashboard Direzione (pagina separata)",
     "Pagina dedicata con KPI aggregati:\n\n• Salute pipeline: totale contatti split NB/Clienti, distribuzione per stadio, velocità funnel (tempo medio per stadio)\n• Performance commerciale: contatti per commerciale con breakdown stadio, rapporto attivi vs fermi/inattivi, contatti mossi nell'ultimo mese\n• Clienti fidelizzati: attivi vs fermi vs inattivi, clienti che l'anno scorso mandavano brief e quest'anno no\n• Settori: distribuzione per industry, settori con molti prospect ma poche conversioni, settori dove ci sono clienti ma zero NB\n• Segnali di allarme: clienti premium fermi, pipeline sbilanciata (tanti prospect, pochi brief), commerciali con portafoglio tutto giallo/rosso\n\nFiltri: per commerciale, settore, stadio, periodo.",
     "Garbarino: \"io voglio sapere cosa sta succedendo, deve uscire subito\". Marco: \"serve una dashboard, senza di questo come fai a controllare?\".",
     "MUST HAVE", "TBC — Quali settori sono prioritari per Noloop? Esiste una classificazione clienti per fascia budget/importanza (premium)? Confermare KPI con Zaghi."),

    ("MA2.5", "Viste per Ruolo", "Dashboard — Logiche business avanzate",
     "Nella dashboard, possibilità di esplorare:\n• Industrie critiche / prioritarie (settori strategici per Noloop)\n• Clienti premium (alto budget o alto potenziale)\n• Concentrazione fatturato (top 5-10 clienti per volume, rischio dipendenza)\n• Tasso conversione tra stadi pipeline",
     "Permette alla direzione di prendere decisioni strategiche basate sui dati, non sull'intuizione.",
     "SHOULD HAVE", "TBC — Definire con Zaghi i criteri di \"premium\" e le industry prioritarie."),

    # MA3 — FILTRI E INDICATORI
    ("section", "MA3 — FILTRI E INDICATORI VISIVI"),
    ("MA3.1", "Filtri e Indicatori", "Filtro per macro-categoria",
     "Filtro per Cliente / New Business nella vista principale dei contatti.",
     "Distinguere immediatamente clienti fidelizzati da prospect.",
     "MUST HAVE", ""),

    ("MA3.2", "Filtri e Indicatori", "Filtro per stadio pipeline",
     "Filtro per stadio (Prospect, Primo Contatto, Follow-up, ecc.) nella vista principale.",
     "Garbarino: \"devo filtrare e mi devono uscire subito\".",
     "MUST HAVE", ""),

    ("MA3.3", "Filtri e Indicatori", "Filtro data con range picker",
     "Il filtro per ultimo contatto/ultima attività diventa un selettore di intervallo temporale (da data X a data Y). Mostra solo i contatti la cui ultima attività cade in quell'intervallo.",
     "Oggi c'è solo l'ordinamento cronologico. Il range picker permette: \"mostrami chi non è stato contattato da gennaio a oggi\".",
     "MUST HAVE", ""),

    ("MA3.4", "Filtri e Indicatori", "Semafori temporali",
     "Indicatore visivo (pallino verde/giallo/rosso) accanto al nome del contatto nella lista e come bordo/badge sulla card Kanban. Calcolato automaticamente in base al tempo dall'ultima attività.\n\nATTENZIONE: il semaforo misura la \"freschezza della relazione\" (quanto tempo è passato). NON replica lo status dell'attività (Positivo/Negativo/In Corso) che misura la \"qualità dell'interazione\". Sono tre dimensioni distinte:\n- Pipeline = dove sei nel funnel\n- Status attività = come è andata l'ultima interazione\n- Semaforo = quanto tempo è passato\n\nPosizionamento: il semaforo va accanto al contatto, non accanto all'attività.",
     "Garbarino vuole identificare a colpo d'occhio situazioni ferme senza cliccare ogni contatto. Il semaforo è l'unico indicatore puramente automatico — l'utente non lo tocca.",
     "MUST HAVE", "TBC — Soglie: proposta NB verde <30gg, giallo 30-90gg, rosso >90gg. Clienti verde <3 mesi, giallo 3-6 mesi, rosso >6 mesi. Confermare con Zaghi."),

    ("MA3.5", "Filtri e Indicatori", "Colonna data ultima attività in tabella",
     "Aggiungere colonna visibile nella tabella principale con la data dell'ultima attività registrata. Senza entrare nella scheda.",
     "Ilia: \"devo rientrare fisicamente dentro ogni contatto per vedere l'ultima volta che l'ho contattato — su 500 contatti è la morte\".",
     "MUST HAVE", ""),

    ("MA3.6", "Filtri e Indicatori", "Colonna stadio pipeline in tabella",
     "Aggiungere colonna visibile nella tabella principale con lo stadio pipeline corrente.",
     "Lettura immediata dello stato senza entrare nella singola scheda.",
     "MUST HAVE", ""),

    # MA4 — GESTIONE AZIENDE
    ("section", "MA4 — GESTIONE AZIENDE EVOLUTA"),
    ("MA4.1", "Gestione Aziende", "Alias aziende",
     "Definire nomi alternativi per la stessa azienda (Pirelli, Pirelli SPA, PIRELLI S.p.A. → tutto sotto \"Pirelli\"). Lo scraper riconduce automaticamente gli alias all'azienda madre. L'utente non deve più correggere manualmente.",
     "Oggi lo scraper ri-separa i contatti sotto aziende diverse quando su LinkedIn il nome è scritto diversamente. Il commerciale corregge, lo scraper sovrascrive.",
     "MUST HAVE", ""),

    ("MA4.2", "Gestione Aziende", "Relazione azienda madre-figlia",
     "Struttura gerarchica a 2 livelli: holding (es. LVMH) → controllate (Fendi, Bulgari, Loro Piana). Vista che aggrega i contatti a livello holding. Ogni azienda figlia mantiene la sua scheda ma è collegata alla madre.",
     "Permette alla direzione di vedere la totalità dei contatti sotto un gruppo, non solo sotto la singola entità giuridica.",
     "MUST HAVE", ""),

    ("MA4.3", "Gestione Aziende", "Vista aziende arricchita",
     "Nella pagina aziende: se madre → mostra figlie collegate + totale contatti aggregato. Se figlia → link alla madre. Navigazione gerarchica.",
     "Navigazione fluida tra holding e controllate senza cercare manualmente.",
     "MUST HAVE", ""),

    # MA5 — NOTIFICHE (EVOLUTIVE)
    ("section", "MA5 — NOTIFICHE E PROATTIVITÀ (EVOLUTIVE)"),
    ("MA5.1", "Notifiche (evolutiva)", "Email giornaliera azioni pending (Sally)",
     "Ogni mattina Sally controlla i lead con azioni in sospeso e manda email al commerciale con recap: cosa fare, con chi, entro quando.",
     "Priorità #1 emersa dal training con i commerciali. Matteo: \"una roba assurda\". Simona: \"sarebbe ottimo\".",
     "NICE TO HAVE", ""),

    ("MA5.2", "Notifiche (evolutiva)", "Sincronizzazione calendario esterno",
     "Sincronizzazione bidirezionale con Outlook/Google Calendar per ricevere promemoria sulle date inserite nelle attività LeadMe.",
     "Richiesta Ilia. Pesante come integrazione (OAuth, gestione token per 8 utenti).",
     "NICE TO HAVE", ""),

    ("MA5.3", "Notifiche (evolutiva)", "Sally AI — Reportistica NLP",
     "Interrogazione in linguaggio naturale sui dati commerciali: \"quanti lead positivi ha Simona nell'ultimo trimestre?\", \"clienti fermi da più di 6 mesi nel settore automotive?\".",
     "Permette alla direzione di interrogare il database senza navigare filtri.",
     "NICE TO HAVE", ""),

    ("MA5.4", "Notifiche (evolutiva)", "Alert clienti dormienti per direzione",
     "Notifica automatica alla direzione quando un cliente premium non riceve brief da X mesi.",
     "Garbarino: \"voglio che mi arrivi un'email che mi dice attenzione, questo cliente è un anno che è lì e non ci ha mai dato niente\".",
     "NICE TO HAVE", ""),

    # MA6 — EXPORT
    ("section", "MA6 — EXPORT DATI"),
    ("MA6.1", "Export Dati", "Export Excel filtrato",
     "Bottone \"Export\" che scarica in .xlsx i dati attualmente visualizzati (con filtri applicati). L'utente filtra → clicca → scarica.",
     "Richiesta Ilia e Marco per analisi esterne e reportistica ad hoc.",
     "SHOULD HAVE", ""),
]

row = 2
for item in reqs:
    if item[0] == "section":
        for col in range(1, 8):
            cell = ws2.cell(row=row, column=col)
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER
            if col == 1:
                cell.value = item[1]
                cell.font = SECTION_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws2.row_dimensions[row].height = 26
        row += 1
        continue

    rid, area, name, desc, why, priority, tbc = item
    vals = [rid, area, name, desc, why, priority, tbc]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

        if col == 6:
            cell.alignment = CENTER
            if val == "MUST HAVE":
                cell.fill = MUST_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="375623")
            elif val == "SHOULD HAVE":
                cell.fill = PatternFill("solid", fgColor="E2D9F3")
                cell.font = Font(name="Arial", size=10, bold=True, color="4A2D7A")
            elif val == "NICE TO HAVE":
                cell.fill = NICE_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="7F6000")

        if col == 7 and "TBC" in val:
            cell.fill = TBC_FILL
            cell.font = Font(name="Arial", size=10, color="B71C1C")

    lines = max(desc.count("\n") + 1, why.count("\n") + 1, tbc.count("\n") + 1 if tbc else 1)
    ws2.row_dimensions[row].height = max(30, min(lines * 15, 200))
    row += 1

# Legend at bottom
row += 1
ws2.cell(row=row, column=1, value="LEGENDA").font = Font(name="Arial", bold=True, size=10)
row += 1

legend = [
    ("MUST HAVE", MUST_FILL, "Requisito essenziale per la deadline del 15 aprile"),
    ("SHOULD HAVE", PatternFill("solid", fgColor="E2D9F3"), "Importante, da includere se budget e tempi lo permettono"),
    ("NICE TO HAVE", NICE_FILL, "Evolutiva futura — fuori scope attuale"),
]
for label, fill, desc in legend:
    c1 = ws2.cell(row=row, column=1, value=label)
    c1.fill = fill
    c1.font = BODY_BOLD
    c1.border = THIN_BORDER
    c1.alignment = CENTER
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c2 = ws2.cell(row=row, column=2)
    c2.fill = fill
    c2.border = THIN_BORDER
    c3 = ws2.cell(row=row, column=3, value=desc)
    c3.font = BODY_FONT
    c3.border = THIN_BORDER
    c3.alignment = WRAP
    ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    for cc in range(4, 8):
        ws2.cell(row=row, column=cc).border = THIN_BORDER
    row += 1

c1 = ws2.cell(row=row, column=1, value="TBC")
c1.fill = TBC_FILL
c1.font = Font(name="Arial", bold=True, size=10, color="B71C1C")
c1.border = THIN_BORDER
c1.alignment = CENTER
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws2.cell(row=row, column=2).fill = TBC_FILL
ws2.cell(row=row, column=2).border = THIN_BORDER
c3 = ws2.cell(row=row, column=3, value="Punto aperto — richiede input/validazione da Laura Zaghi")
c3.font = BODY_FONT
c3.border = THIN_BORDER
c3.alignment = WRAP
ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
for cc in range(4, 8):
    ws2.cell(row=row, column=cc).border = THIN_BORDER

wb.save("/home/claude/LeadMe_Evolution_RequirementLog.xlsx")
print("Done")
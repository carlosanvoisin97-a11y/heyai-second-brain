import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Requirement Log Wave 2"

# === STYLES ===
DARK_BG = PatternFill("solid", fgColor="1B2537")
SECTION_BG = PatternFill("solid", fgColor="2C3E5A")
SUBSECTION_BG = PatternFill("solid", fgColor="3A4F6F")
ROW_BG = PatternFill("solid", fgColor="F8FAFC")
ALT_BG = PatternFill("solid", fgColor="FFFFFF")
INCL_BG = PatternFill("solid", fgColor="E2EFDA")
OPT_BG = PatternFill("solid", fgColor="FFF2CC")
EXCL_BG = PatternFill("solid", fgColor="FCE4EC")
NOTE_BG = PatternFill("solid", fgColor="E8EAF6")
TBD_BG = PatternFill("solid", fgColor="FFF8E1")
REF_BG = PatternFill("solid", fgColor="E3F2FD")

WHITE_F = Font(name="Arial", bold=True, color="FFFFFF", size=10)
WHITE_SM = Font(name="Arial", color="FFFFFF", size=9)
DARK_F = Font(name="Arial", size=10, color="1B2537")
DARK_B = Font(name="Arial", size=10, color="1B2537", bold=True)
GRAY_F = Font(name="Arial", size=9, color="6B7280")
RED_F = Font(name="Arial", size=9, color="B71C1C")
GREEN_F = Font(name="Arial", size=9, bold=True, color="2E7D32")
ORANGE_F = Font(name="Arial", size=9, bold=True, color="E65100")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN = Border(
    left=Side("thin", color="D0D5DD"), right=Side("thin", color="D0D5DD"),
    top=Side("thin", color="D0D5DD"), bottom=Side("thin", color="D0D5DD")
)

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

# === HEADER ===
headers = ["ID", "Sotto-area", "Requisito", "Perimetro", "Note / Vincoli", "Costo (€)", "Vendita (€)"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.font = WHITE_F
    cell.fill = DARK_BG
    cell.alignment = CENTER
    cell.border = THIN
ws.row_dimensions[1].height = 24
ws.freeze_panes = "A2"

row = 2

def section(r, title, cost=None, sale=None):
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.fill = SECTION_BG
        cell.font = WHITE_F
        cell.border = THIN
        cell.alignment = Alignment(vertical="center")
    ws.cell(r, 1, title)
    if cost: ws.cell(r, 6, cost).number_format = '#,##0'
    if sale: ws.cell(r, 7, sale).number_format = '#,##0'
    ws.row_dimensions[r].height = 22
    return r + 1

def subsection(r, title):
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.fill = SUBSECTION_BG
        cell.font = WHITE_SM
        cell.border = THIN
    ws.cell(r, 1, title)
    ws.row_dimensions[r].height = 18
    return r + 1

def req(r, rid, area, name, scope, note="", cost=None, sale=None):
    bg = ROW_BG if r % 2 == 0 else ALT_BG
    scope_bg = INCL_BG if scope == "INCLUSO" else OPT_BG if scope == "OPZIONALE" else EXCL_BG if scope == "ESCLUSO" else NOTE_BG if scope == "NOTA" else TBD_BG if scope == "DA DEFINIRE" else REF_BG
    scope_f = GREEN_F if scope == "INCLUSO" else ORANGE_F if scope == "OPZIONALE" else RED_F
    
    vals = [rid, area, name, scope, note, cost, sale]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = DARK_F if c < 4 else scope_f if c == 4 else GRAY_F
        cell.fill = scope_bg if c == 4 else bg
        cell.alignment = WRAP
        cell.border = THIN
        if c >= 6 and v:
            cell.number_format = '#,##0'
            cell.alignment = CENTER
    
    lines = max((name or "").count("\n") + 1, (note or "").count("\n") + 1)
    ws.row_dimensions[r].height = max(28, min(lines * 14, 160))
    return r + 1

def ref_section(r, title, note):
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.fill = REF_BG
        cell.font = Font(name="Arial", size=10, color="1565C0", italic=True)
        cell.border = THIN
    ws.cell(r, 1, title)
    ws.cell(r, 5, note)
    ws.row_dimensions[r].height = 22
    return r + 1

# =============================================
# MODULI CON REQUIREMENT LOG SEPARATO
# =============================================
row = ref_section(row, "▸ LEADME EVOLUTION", "Requirement log separato (32 requisiti, 22 MUST, deadline 15/04). Vedi file dedicato.")
row = ref_section(row, "▸ VENUE FINDER EVOLUTION", "Requirement log separato. Sviluppo in corso (Manu). Vedi file dedicato.")
row = ref_section(row, "▸ MINUTES EVOLUTION", "Requirement log separato. Registrazione call esterne = priorità. Vedi file dedicato.")

# =============================================
# FLOW
# =============================================
row = section(row, "▌ FL — FLOW — TASK MANAGEMENT", 40000, 50000)

row = subsection(row, "   FL-A — Core Task Management")

row = req(row, "FL.1", "Task", 
    "Ricezione task strutturati da Minutes: assegnatario, scadenza, note, priorità. Client lead rivede, EDITA e conferma con un click → pushati su Flow",
    "INCLUSO",
    "I task da Minutes devono poter essere editati prima dell'invio a Flow (nuovo req da recap strategico). Avviene in due momenti: post-brainstorming e post-accettazione.")

row = req(row, "FL.2", "Task",
    "Vista Kanban: Pending / In Progress / Completed / Deferred. Vista per persona e per commessa",
    "INCLUSO",
    "Kanban standard. Componente riutilizzabile per vista clienti su LeadMe/CRM.")

row = req(row, "FL.3", "Task",
    "Possibilità vista Gantt per commessa",
    "INCLUSO",
    "Vista alternativa al Kanban per commesse complesse.")

row = req(row, "FL.4", "Task",
    "Task liberi + task tipizzati con menù dropdown. Procedure interne allegabili a task tipizzati",
    "INCLUSO",
    "Tipi task da definire con il Cliente in fase di configurazione.")

row = req(row, "FL.5", "Task",
    "Upload file nei task: immagini, foto, documentazione, fatture, contratti",
    "INCLUSO",
    "File poi visibili anche su Link.")

row = req(row, "FL.6", "Input",
    "Flow accetta input non strutturati: audio, foto, copia-incolla email → PROPONE task ma utente deve confermare. Nessuna azione autonoma",
    "INCLUSO",
    "Da recap strategico: Flow ha potere di PROPOSTA con CONFERMA, non potere di azione. L'utente valida sempre prima dell'esecuzione.")

row = subsection(row, "   FL-B — Sally in Flow")

row = req(row, "FL.7", "Sally",
    "Interazione Sally vocale/testo: aggiornamento task, caricamento file, cambio stato via linguaggio naturale",
    "INCLUSO",
    "Es. \"Ho fatto il sopralluogo, allego foto\" → Sally identifica task, aggiorna. Costi API LLM a carico Cliente.")

row = req(row, "FL.8", "Sally",
    "Upload contratti con estrazione scadenze: importo, condizioni pagamento, deadline operative → diventano reminder automatici su Flow",
    "INCLUSO",
    "Skill Sally quotata in Flow. Alimenta reminder per rooming list, biglietti, cancellazioni.")

row = req(row, "FL.9", "Sally",
    "Reportistica NLP Flow: \"quali task sono in ritardo sulla commessa X?\", \"chi ha il carico più alto?\"",
    "INCLUSO",
    "Skill Sally quotata in Flow.")

row = req(row, "FL.10", "Sally",
    "Validazione procedurale AI: Sally confronta avanzamento con procedure aziendali, blocca/segnala se non rispettate",
    "OPZIONALE",
    "⚠️ ADD-ON COSTOSO. Richiede knowledge base procedure completa.")

row = subsection(row, "   FL-C — Chiusura & Integrazione")

row = req(row, "FL.11", "Validazione",
    "Blocco chiusura commessa se dati mancanti (fatture, documenti obbligatori)",
    "INCLUSO",
    "Regole di validazione configurabili da Admin.")

row = req(row, "FL.12", "Integrazione",
    "Push dati aggregati su Link al completamento di tutti i task",
    "INCLUSO",
    "Auto-population gestionale.")

row = req(row, "FL.13", "Budget",
    "Transito Budget 2 (previsionale) e Budget 3 (consuntivo) verso Link, con verifiche completezza",
    "INCLUSO",
    "Meccanismo esatto da definire in fase di design. I task relativi al budget possono alimentare il budget di Link.")

row = req(row, "FL.14", "Workflow",
    "Rispetto workflow approvazioni BC: quando un task genera richiesta di pagamento, il sistema facilita la doppia approvazione (Resp. Reparto → Resp. Controllo/Direzione)",
    "INCLUSO",
    "Flow NON sostituisce l'approvazione ma la facilita (alert, pre-compilazione). Tempistiche BC: standard lunedì h15, urgente h12 stesso giorno.")

row = subsection(row, "   FL — ⚠️ NOTA ARCHITETTURALE")

row = req(row, "FL.N", "ARCHITETTURA",
    "PUNTO APERTO: Flow come applicativo separato o integrato nell'ecosistema Minutes+LeadMe?\n\n"
    "Opzione A: Flow è app separata (posizione Federico — IP rivendibile)\n"
    "Opzione B: Le funzionalità di Flow vengono assorbite dalla combo Minutes (raccoglie input, estrae task) + LeadMe/CRM (vista clienti a funnel con task sotto le commesse) + Sally (orchestra)\n\n"
    "In entrambi i casi: riutilizzare componenti (pipeline Kanban, vista task) tra i moduli. Federico deve fare proposta tecnica.",
    "DA DEFINIRE",
    "Da recap strategico: il cliente non vuole comprare Flow come app separata. Dobbiamo trovare il modo di infilarlo nell'ecosistema senza un applicativo in più. Questa decisione impatta roadmap e Gantt per Donzelli.")

# =============================================
# LINK
# =============================================
row = section(row, "▌ LK — LINK — GESTIONALE CENTRALIZZATO", 40000, 50000)

row = subsection(row, "   LK-N — ⚠️ NOTA PERIMETRO")

row = req(row, "LK.0", "PERIMETRO",
    "Link = centralizzatore dati + integratore sistemi + reportistica (Progress digitale).\n"
    "NO ciclo attivo, NO ciclo passivo, NO gestione fornitori, NO anagrafiche fornitori, NO fatturazione.\n"
    "Tutto su Business Central.",
    "NOTA",
    "Se BC non espone API → scope si riduce ulteriormente a solo Progress digitale. In quel caso il prezzo resta invariato ma si aggiungono altre feature per compensare (da recap strategico).")

row = subsection(row, "   LK-A — Commesse & Budget")

row = req(row, "LK.1", "Commesse",
    "Creazione commessa: riceve da Minutes (primario) o LeadMe. Struttura: cliente, evento, date, budget iniziale, margine atteso, regime fiscale, tipologia fatturazione",
    "INCLUSO",
    "Guscio vuoto che si riempie nel tempo con dati da altri moduli.")

row = req(row, "LK.2", "Commesse",
    "Gatekeeper con notifica: alla creazione commessa il Resp. Controllo riceve notifica con dati pre-compilati. Conferma, corregge o respinge prima del push verso BC",
    "INCLUSO",
    "Preserva il ruolo attuale del Resp. Controllo nella validazione del regime fiscale. Vincolo BC: non si possono avere due regimi fiscali in una commessa.")

row = req(row, "LK.3", "Budget",
    "Budget digitale a 3 stadi + snapshot automatico:\n"
    "Stadio 1 — In gara (macro dal brief/Minutes)\n"
    "Stadio 2 — Pre-chiusura (costi contrattati)\n"
    "Stadio 3 — Consuntivo (costi reali con fatture)\n"
    "Snapshot immutabile ad ogni passaggio. Facing temporale consultabile dalla direzione.",
    "INCLUSO",
    "Il sistema manda alert a responsabili per aggiornare pre-chiusura prima della partenza. NB: il cliente NON vuole fare il budget digitale su Link, vuole continuare con Excel. I 3 stadi entrano su Link nei momenti definiti, non si replica Excel.")

row = req(row, "LK.4", "Budget",
    "Approvazione Budget 2 previsionale per clienti importanti: workflow con gate Garbarino",
    "INCLUSO",
    "Workflow approvazione configurabile.")

row = subsection(row, "   LK-B — Piano di Produzione Digitale (PDP)")

row = req(row, "LK.5", "PDP",
    "Piano di Produzione Digitale: strumento di lavoro quotidiano per operativi e producer.\n"
    "Template intelligenti pre-popolati basati su storico commesse.\n"
    "Import template cliente (drag & drop). Flessibilità totale (colonne aggiungibili, filtrabili).\n"
    "Associazione fatture passive alle singole voci.\n"
    "Suggerimento fornitori da BC per commesse simili (solo nome, NO costo).",
    "INCLUSO",
    "⚠️ ARCHITETTURA: PDP può vivere dentro Flow (dove la gente lavora) o dentro Link (dove convergono dati economici). Da decidere. Quotare entrambe le opzioni se le stime differiscono.")

row = req(row, "LK.6", "PDP",
    "Associazione fatture passive — DUE CANALI:\n"
    "Canale 1 — Cassetto Fiscale (SDI): cross-check con BC tramite codice commessa\n"
    "Canale 2 — Non-SDI (fatture estere, regime minimi): verifica upload PDF nella cartella con naming standard\n"
    "Distinguere proforma (archivio) da fatture finali (fiscali).",
    "INCLUSO",
    "Fatture multi-commessa: in BC il Nr. Commessa va a livello di riga, non di testata.")

row = req(row, "LK.7", "PDP",
    "Template fatturazione attiva: operativo compila su Link (stessa UX dell'Excel attuale: PO, importi, codici IVA). Push a BC via API. Resp. Amministrativo conferma in BC",
    "INCLUSO",
    "Elimina passaggio email a fatturazioneatttiva@noloop.eu. Codici IVA: ART-000004 (22%), ART-000005 (10%), ART-000006 (ESC 15), ART-000007 (Non Imp 9).")

row = req(row, "LK.8", "PDP",
    "Blocco chiusura PDP: non può passare in \"completed\" finché tutte le voci non hanno fattura associata, template fatturazione compilati, check quadratura PDP↔BC superato",
    "INCLUSO",
    "Regola ferrea: niente chiusura senza fatture. Gestione spese Revolut: cross-check con spese non associate a COM.")

row = subsection(row, "   LK-C — Integrazioni Business Central")

row = req(row, "LK.9", "Integrazione",
    "Push commessa verso BC via API Microsoft (OAuth). Link → BC",
    "INCLUSO",
    "API BC confermata disponibile (E.R.). Formato codice: AALNNNNN. Testare su commessa standard ASAP. Eventuali limitazioni API BC non imputabili al fornitore.")

row = req(row, "LK.10", "Integrazione",
    "Integrazione cash flow BC:\n"
    "Ricavi → Righe Pianificazione Commessa (task RICAVI)\n"
    "Costi → Richieste di Pagamento in stato Pianificato\n"
    "Controllo residuo automatico: se importo supera pianificato, collegamento non possibile",
    "INCLUSO",
    "Regole BC: ricavi al netto IVA, costi IVA inclusa. Rischedulazione residui 3 cicli × 30gg. Valuta estera: cash flow in euro, richiesta pagamento in valuta fornitore.")

row = req(row, "LK.11", "Integrazione",
    "Ricezione dati da Flow: task completati, costi confermati",
    "INCLUSO",
    "Auto-population.")

row = req(row, "LK.12", "Integrazione",
    "Feedback bidirezionale verso LeadMe: brief ricevuto, gara vinta, progetto chiuso",
    "INCLUSO",
    "Tiene aggiornato status commerciale.")

row = req(row, "LK.13", "Integrazione",
    "Gestione GARE: se gara persa → status \"non confermata\" → chiusura COM su BC. Se confermata, comunicare variazioni (budget/località/data)",
    "INCLUSO",
    "Nuovo requisito dal brief, mancante dall'Excel precedente.")

row = subsection(row, "   LK-D — Dashboard & Reportistica")

row = req(row, "LK.14", "Dashboard",
    "Dashboard Garbarino (sostituisce Progress Excel): status commessa per BU/commerciale/tipo cliente, budget per reparto, marginalità, facing a 3 stadi, comparazione previsto vs consuntivo",
    "INCLUSO",
    "Auto-alimentata dai dati PDP + task completati Flow.")

row = req(row, "LK.15", "Dashboard",
    "Analisi per cliente, BU, tipo progetto. Carichi di lavoro per persona",
    "INCLUSO")

row = req(row, "LK.16", "Sally",
    "Reportistica NLP Link: interrogazione linguaggio naturale (\"qual è il margine sulla commessa X?\")",
    "INCLUSO",
    "Skill Sally quotata in Link. Costi API LLM a carico Cliente.")

row = subsection(row, "   LK — Add-on")

row = req(row, "LK.A", "Add-on",
    "Integrazione Revolut: cross-check spese carta vs commesse. Alert spese non associate a codice COM",
    "OPZIONALE",
    "⚠️ ADD-ON separato. Cambio provider in corso. Verifica sintassi AAXNNNNN. Limite €500/transazione.")

# =============================================
# SALLY
# =============================================
row = section(row, "▌ SL — SALLY — AGENTE VIRTUALE", 30000, 40000)

row = subsection(row, "   SL-A — Infrastruttura Core")

row = req(row, "SL.1", "Core",
    "Backend dedicato standalone: backend e frontend separati da Link. Funzioni chat migrate da Minutes",
    "INCLUSO",
    "Deploy indipendente. Cambio prompt da env.")

row = req(row, "SL.2", "Core",
    "PWA installabile su telefono (stile ChatGPT): voce + testo",
    "INCLUSO",
    "NO app nativa. Web-based. Compatibilità iOS/Android via browser.")

row = req(row, "SL.3", "Core",
    "Knowledge base procedure aziendali Noloop: Sally risponde come \"collega esperto\"\n"
    "Include: richieste di pagamento (3 casistiche ritenute), fatturazione attiva, gestione Revolut, fatture estere, apertura commessa, cash flow",
    "INCLUSO",
    "~2.000-3.000 documenti. Sincronizzazione notturna con cartelle sorgente. Riduce ~120 email/settimana a Calajò per supporto procedurale.")

row = req(row, "SL.4", "Core",
    "Notifiche proattive (Livello 2): cron job giornaliero.\n"
    "Check lead dormienti su LeadMe, alert scadenze contrattuali su Flow, reminder pre-chiusura commessa, segnalazione anomalie",
    "INCLUSO",
    "Email + notifica in-app. Sally prende iniziativa.")

row = subsection(row, "   SL-B — Architettura Skill")

row = req(row, "SL.5", "Skill",
    "Architettura a plugin/skill modulari: funzionalità AI specifiche per LeadMe, Flow, Link",
    "INCLUSO",
    "Le singole skill sono quotate nei rispettivi moduli (LM.7, LM.8, FL.7, FL.9, LK.16).")

row = req(row, "SL.6", "Skill",
    "Interrogazione NLP in-app (Livello 1): Sally presente come assistente in ogni applicativo. Lettura dati + risposta. Nessuna azione esecutiva",
    "INCLUSO",
    "Quotata nei rispettivi moduli.")

row = req(row, "SL.7", "Skill",
    "Azioni esecutive cross-modulo (Livello 2): Sally può creare/aggiornare task su Flow, aggiornare status su LeadMe, inserire voci nel PDP.\n"
    "L'utente CONFERMA l'azione proposta con un click prima dell'esecuzione",
    "INCLUSO",
    "Richiede integrazioni API con ciascun modulo. Quotare il costo di integrazione per modulo.")

row = req(row, "SL.8", "Skill",
    "Email-to-action / Audio-to-action (Livello 2): utente inoltra email o invia nota vocale/foto → Sally analizza, propone azione e destinazione → utente conferma → Sally esegue",
    "INCLUSO",
    "Livello 2: richiede potere esecutivo su Flow, Link, LeadMe.")

row = req(row, "SL.9", "Skill",
    "Input non strutturati → PDP (Livello 2): foto scontrini, email fornitori, note vocali → Sally formalizza e inserisce nella riga corretta del PDP",
    "INCLUSO",
    "Estrae da contratti: importo, condizioni pagamento, scadenze.")

row = req(row, "SL.10", "Skill",
    "Compilazione documenti e portali fornitori: pre-compilazione autocertificazioni, DUVRI, form Jaggaer/Ariba con dati dalla KB.\n"
    "Fase 1: Word e PDF compilabili. Fase 2: PDF immagine con OCR",
    "INCLUSO",
    "~80 portali fornitori. Scadenze periodiche: DURC ogni 6 mesi, visure, certificazioni. Riduce 5-10h/settimana di Marco M. + Davide.")

row = req(row, "SL.11", "Skill",
    "Memoria condivisa tra utenti: Sally ha visibilità cross-utente per performance migliori e risposte contestualizzate",
    "INCLUSO",
    "⚠️ RISCHIO: se il sistema allucina e tira fuori informazione che non dovrebbe (insulto a dipendente, informazioni pagamento, ecc.) può creare crisi interna. Federico la vuole per performance. Presentare al cliente con nota esplicita sul rischio di allucinazione.")

row = subsection(row, "   SL — Add-on")

row = req(row, "SL.A", "Add-on",
    "Voice mode live (conversazione vocale real-time stile GPT Voice)",
    "OPZIONALE",
    "⚠️ PREMIUM. Costo token significativo. Da valutare sostenibilità con Cliente.")

# =============================================
# ONSITE / CROWD
# =============================================
row = section(row, "▌ ON — ONSITE — APP PARTECIPANTI (Crowd Estensione)", 18000, 25000)

row = subsection(row, "   ON-A — Funzionalità")

row = req(row, "ON.1", "App",
    "PWA Social networking partecipanti: programma personalizzato, chatbot evento",
    "INCLUSO",
    "Estensione di Crowd per il durante-evento.")

row = req(row, "ON.2", "Integrazione",
    "Collegamento evento Crowd a commessa Link",
    "INCLUSO",
    "Già previsto nell'architettura Crowd.")

row = req(row, "ON.3", "Pagamenti",
    "Gestione pagamenti integrata in Crowd",
    "INCLUSO",
    "Gateway pagamento da definire. Da aggiungere a stima Crowd attuale.")

row = req(row, "ON.4", "Feedback",
    "Survey e feedback strutturati: form personalizzabili durante e dopo evento. NPS, raccolta input qualitativi. Dati aggregati per report post-evento",
    "INCLUSO",
    "Nuovo requisito dal brief, mancante dall'Excel precedente.")

# =============================================
# ESCLUSIONI
# =============================================
row = section(row, "▌ EX — ESCLUSIONI GENERALI WAVE 2")

row = req(row, "EX.1", "Esclusione", "Ciclo attivo, ciclo passivo, gestione fornitori, anagrafiche su Link (resta su BC)", "ESCLUSO")
row = req(row, "EX.2", "Esclusione", "Replica funzionalità Excel per budget (i budget restano su Excel, entrano su Link nei momenti definiti)", "ESCLUSO",
    "Il cliente insiste: non vuole fare il budget digitale su Link.")
row = req(row, "EX.3", "Esclusione", "Pitch — Modulo Creativo (Deep Search + Playground)", "ESCLUSO",
    "Cliente non lo vuole. Rimosso dal perimetro.")
row = req(row, "EX.4", "Esclusione", "Formazione operatori, supporto on-site, change management", "ESCLUSO", "Quotabile separatamente.")
row = req(row, "EX.5", "Esclusione", "Manutenzione evolutiva e supporto post-rilascio", "ESCLUSO", "Quotabile con contratto annuale.")
row = req(row, "EX.6", "Esclusione", "Hosting e infrastruttura ricorrente", "ESCLUSO", "A carico Cliente.")
row = req(row, "EX.7", "Esclusione", "Costi API LLM (OpenAI/Anthropic) per tutte le funzionalità Sally", "ESCLUSO", "Consumo token a carico Cliente.")

# =============================================
# TOTALI
# =============================================
row += 1
ws.cell(row, 1, "SUBTOTALE — Flow + Link + Sally + OnSite").font = DARK_B
ws.cell(row, 6, 128000).font = DARK_B
ws.cell(row, 6).number_format = '#,##0'
ws.cell(row, 7, 165000).font = DARK_B
ws.cell(row, 7).number_format = '#,##0'
row += 1
ws.cell(row, 1, "NB: LeadMe, Venue Finder, Minutes hanno stime separate nei rispettivi req log").font = GRAY_F
row += 1
ws.cell(row, 1, "NB: Costi cella F sono le stime di Federico/Simone sull'Excel precedente — DA AGGIORNARE con i requisiti mancanti").font = RED_F
row += 1
ws.cell(row, 1, "NB: La decisione architetturale su Flow impatta tutte le stime. Non finalizzare prima della proposta di Federico").font = RED_F

wb.save("/home/claude/Wave2_ReqLog_Completo.xlsx")
print("Done")